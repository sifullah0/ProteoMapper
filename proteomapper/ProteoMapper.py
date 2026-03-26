#final version with detailed timing and optimizations
import subprocess
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
import numpy as np
import re
import os
import platform
import tempfile
import multiprocessing
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Alignment, Border, Side, Font
from openpyxl.comments import Comment
from collections import Counter, defaultdict
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Custom Exceptions
class HeaderNotFoundError(Exception):
    """Raised when header row cannot be detected in Excel file"""
    pass

class FileReadError(Exception):
    """Raised when Excel file cannot be read"""
    pass

class RegexFileError(Exception):
    """Raised when regex patterns file cannot be read"""
    pass

class PositionsFileError(Exception):
    """Raised when positions file cannot be read"""
    pass

class HMMERError(Exception):
    """Raised when HMMER execution fails"""
    pass

# Helper Functions

# ----------  motif parallel helpers  ----------
import multiprocessing as mp
from collections import Counter, defaultdict
from openpyxl.styles import PatternFill

def _match_chunk(params):
    """
    Runs regex matching on one slice of the dataframe.
    Returns a *serialisable* dict with everything the main process
    needs to reconstruct the highlights + MDCS inputs.
    """
    chunk_df, patterns, start_row, leading_columns = params
    ranges = []
    summary = defaultdict(int)
    pattern_ranges = defaultdict(list)

    # NEW: collect motif hits in GAPLESS (match_seq) coordinates for MDCS.
    # Each item: (sequence_index_1based, motif_pattern, start_no_gap_1based, end_no_gap_1based_inclusive)
    motif_hits = []

    # Pre-compile regexes
    compiled_patterns = [(re.compile(pattern, re.IGNORECASE), label)
                         for pattern, label in patterns]

    for idx, (inner_idx, row) in enumerate(chunk_df.iterrows()):
        cleaned = row['protein sequences_cleaned']
        match_seq = row['protein sequences_match'].upper()

        # OPTIMIZATION: Pre-compute position map once per sequence
        # Maps indices in match_seq (no gaps) to indices in cleaned_seq (with gaps)
        pos_map = []
        for i, ch in enumerate(cleaned):
            if ch not in {" ", "-"}:
                pos_map.append(i)

        # Skip empty sequences
        if not pos_map:
            continue

        # sequence index as used elsewhere in the pipeline (1..N)
        sequence_index = (start_row + idx) - 1  # excel row starts at 2; sequence index starts at 1

        for cre, label in compiled_patterns:
            key = label or cre.pattern

            # Find all matches in the gapless sequence
            for m in cre.finditer(match_seq):
                # gapless coords (1-based, inclusive)
                m_start_ng = m.start() + 1
                m_end_ng = m.end()  # since end() is exclusive, this is inclusive end in 1-based
                motif_hits.append((sequence_index, key, m_start_ng, m_end_ng))

                # map to cleaned (with gaps) for highlighting/summary
                start_cleaned = pos_map[m.start()]
                end_cleaned = pos_map[min(m.end() - 1, len(pos_map) - 1)]

                abs_row = start_row + idx
                ranges.append((abs_row, start_cleaned, end_cleaned))
                summary[key] += 1
                pattern_ranges[key].append((start_cleaned + 1, end_cleaned + 1))

    return {"ranges": ranges, "summary": summary, "pattern_ranges": pattern_ranges, "motif_hits": motif_hits}

def detect_header_row(df):
    """Detect the header row in the dataframe"""
    for index, row in df.iterrows():
        row_normalized = row.astype(str).str.replace(r"\s+", " ", regex=True).str.strip().str.lower()
        if 'protein sequences' in row_normalized.values:
            # Check if any valid identifier column exists
            valid_identifiers = ['gene name', 'gene id', 'protein name', 'protein id']
            if any(identifier in row_normalized.values for identifier in valid_identifiers):
                return index
    return None

def read_fasta_file(fasta_file):
    """Read FASTA file and convert to DataFrame format compatible with Excel pipeline"""
    read_start = time.time()
    
    sequences = []
    current_id = None
    current_name = None
    current_seq = []
    
    try:
        with open(fasta_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line.startswith('>'):
                    # Save previous sequence if exists
                    if current_id is not None:
                        sequences.append({
                            'Protein ID': current_id,
                            'Protein Name': current_name,
                            'Protein Sequences': ''.join(current_seq)
                        })
                    
                    # Parse header line: >ID Name or >ID|Name or >ID
                    header = line[1:].strip()
                    parts = header.split(None, 1)  # Split on first whitespace
                    
                    if len(parts) >= 1:
                        # Handle formats like "sp|P12345|NAME" or just "P12345"
                        if '|' in parts[0]:
                            id_parts = parts[0].split('|')
                            current_id = id_parts[1] if len(id_parts) > 1 else id_parts[0]
                        else:
                            current_id = parts[0]
                        
                        # Use the rest as name, or use ID if no name present
                        current_name = parts[1] if len(parts) > 1 else current_id
                    else:
                        current_id = header
                        current_name = header
                    
                    current_seq = []
                elif line:
                    current_seq.append(line)
            
            # Save last sequence
            if current_id is not None:
                sequences.append({
                    'Protein ID': current_id,
                    'Protein Name': current_name,
                    'Protein Sequences': ''.join(current_seq)
                })
        
        if not sequences:
            raise FileReadError("No sequences found in FASTA file")
        
        # Create DataFrame
        df = pd.DataFrame(sequences)
        
        read_time = time.time() - read_start
        print(f"  ⏱️  FASTA read time: {read_time:.3f} seconds")
        print(f"  📊 Loaded {len(df)} sequences from FASTA file")
        
        return df
        
    except Exception as e:
        raise FileReadError(f"Failed to read FASTA file: {str(e)}")

def read_and_preprocess_data(input_excel_file):
    """Read Excel file and preprocess data"""
    read_start = time.time()
    
    try:
        df_raw = pd.read_excel(input_excel_file, sheet_name=0, header=None, dtype=str)
    except Exception as e:
        raise FileReadError(f"Failed to read Excel file: {str(e)}")
    
    # Find the header row dynamically
    header_row_index = detect_header_row(df_raw)
    if header_row_index is None:
        raise HeaderNotFoundError("Could not find the correct header row. Expected columns: 'Protein Sequences' and at least one identifier column ('Gene Name', 'Gene ID', 'Protein Name', or 'Protein ID')")
    
    # Read the file again using the detected header row
    try:
        df = pd.read_excel(input_excel_file, sheet_name=0, skiprows=header_row_index, dtype=str)
    except Exception as e:
        raise FileReadError(f"Failed to read Excel file with detected header: {str(e)}")
    
    read_time = time.time() - read_start
    print(f"  ⏱️  Excel read time: {read_time:.3f} seconds")
    
    return df

def clean_and_process_sequences(df):
    """Clean and process protein sequences"""
    # Store original column names for identifier columns before normalization
    original_headers = {}
    for col in df.columns:
        col_normalized = str(col).replace(r"\s+", " ").strip().lower()
        col_normalized = ' '.join(col_normalized.split())  # Remove extra spaces
        if col_normalized in ['gene name', 'gene id', 'protein name', 'protein id']:
            original_headers[col_normalized] = col
    
    # Normalize column names
    df.columns = df.columns.str.replace(r"\s+", " ", regex=True).str.strip().str.lower()
    
    # Store the original headers in the dataframe for later use
    df.attrs['original_headers'] = original_headers
    
    # Process sequences
    df['protein sequences'] = df['protein sequences'].fillna("").astype(str)
    df['has_leading_gt'] = df['protein sequences'].str.startswith('>')
    df['Protein sequences'] = df['protein sequences'].apply(lambda x: x[1:] if x.startswith('>') else x)
    
    # Keep original case and gaps
    df['protein sequences_cleaned'] = df['Protein sequences'].str.replace(r"[^ACDEFGHIKLMNPQRSTVWYacdefghiklmnpqrstvwy\-]", "", regex=True)
    
    # Uppercase version for matching (no gaps)
    df['protein sequences_match'] = df['Protein sequences'].str.upper().str.replace(r"[^ACDEFGHIKLMNPQRSTVWY]", "", regex=True)
    return df
                
def prepare_final_dataframe(df):
    """Split sequences and prepare final dataframe"""
    # Split the cleaned sequences into characters
    split_sequences = df['protein sequences_cleaned'].apply(lambda seq: list(seq)).tolist()
    max_length = max(map(len, split_sequences)) if split_sequences else 0
    
    # Create sequence columns
    sequence_columns = [str(i + 1) for i in range(max_length)]
    df_sequence_split = pd.DataFrame(split_sequences, columns=sequence_columns)
    
    # Prepare filtered dataframe - check for all possible identifier columns
    identifier_cols = ['gene name', 'gene id', 'protein name', 'protein id']
    available_cols = [col for col in identifier_cols if col in df.columns]
    
    df_filtered = (
        df[available_cols]
        .replace("", np.nan)
        .dropna(axis=1, how='all')
    )
    
    # Get original headers if available
    original_headers = df.attrs.get('original_headers', {})
    
    # Rename columns - use original headers if available, otherwise use standardized format
    rename_map = {}
    if 'gene name' in df_filtered.columns:
        rename_map['gene name'] = original_headers.get('gene name', 'Gene Name')
    if 'gene id' in df_filtered.columns:
        rename_map['gene id'] = original_headers.get('gene id', 'Gene ID')
    if 'protein name' in df_filtered.columns:
        rename_map['protein name'] = original_headers.get('protein name', 'Protein Name')
    if 'protein id' in df_filtered.columns:
        rename_map['protein id'] = original_headers.get('protein id', 'Protein ID')
    df_filtered = df_filtered.rename(columns=rename_map)
    
    # Merge with the split sequence matrix
    df_final = pd.concat([df_filtered, df_sequence_split], axis=1)
    
    return df_final, df_filtered.shape[1]
    
def setup_excel_styles():
    """Define and return style objects"""
    return {
        'sky_blue_fill': PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid"),
        'green_fill': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        'orange_fill': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        'red_thick_border': Border(
            left=Side(border_style="thick", color="FF0000"),
            right=Side(border_style="thick", color="FF0000"),
            top=Side(border_style="thick", color="FF0000"),
            bottom=Side(border_style="thick", color="FF0000")
        ),
        'bold_font': Font(bold=True),
        'center_alignment': Alignment(horizontal='center', vertical='center'),
        'left_alignment': Alignment(horizontal='left', vertical='center')
    }

def adjust_column_widths(ws):
    """Adjust column widths based on content (restored from previous version for better fitting)"""
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 50)

def apply_bulk_alignment(ws, alignment_style, min_row=2, max_row=None, min_col=1, max_col=None):
    """OPTIMIZED: Apply alignment to entire range at once"""
    if max_row is None:
        max_row = ws.max_row
    if max_col is None:
        max_col = ws.max_column
    
    # Apply alignment to each cell in range
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.alignment = alignment_style
            
def load_patterns(regex_input_file):
    """Load patterns from input file"""
    pattern_label_pairs = []
    try:
        with open(regex_input_file, "r") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                if "::" in line:
                    pattern, label = line.split("::", 1)
                else:
                    pattern, label = line, ""
                pattern_label_pairs.append((pattern.strip(), label.strip()))
    except Exception as e:
        raise RegexFileError(f"Failed to read regex patterns file: {str(e)}")
    return pattern_label_pairs

def _apply_pattern_highlights_ST(ws, df, pattern_label_pairs, leading_columns_count, styles):
    """Apply pattern matching highlights to worksheet"""
    highlight_start = time.time()

    highlight_ranges = []
    print_ranges = []
    match_summary = defaultdict(int)
    pattern_print_ranges = defaultdict(list)
    any_matches_found = False

    # NEW: collect motif hits in GAPLESS coordinates for MDCS
    motif_hits = []

    for pattern, label in pattern_label_pairs:
        try:
            regex = re.compile(pattern, re.IGNORECASE)
        except re.error as e:
            raise ValueError(f"Invalid regular expression '{pattern}': {str(e)}")

        for row_idx, (cleaned_seq, match_seq) in enumerate(zip(df['protein sequences_cleaned'], df['protein sequences_match']), start=2):
            match_seq_upper = match_seq.upper()
            sequence_index = row_idx - 1  # 1-based sequence index

            for match in regex.finditer(match_seq_upper):
                any_matches_found = True

                # gapless coords (1-based, inclusive)
                key = label or pattern
                m_start_ng = match.start() + 1
                m_end_ng = match.end()  # end() is exclusive -> inclusive end in 1-based
                motif_hits.append((sequence_index, key, m_start_ng, m_end_ng))

                # map to cleaned (with gaps) for highlighting/summary
                start_idx, end_idx = match.start(), match.end() - 1
                char_count = 0
                positions = []
                for i, char in enumerate(cleaned_seq):
                    if char not in {" ", "-"}:
                        if start_idx <= char_count <= end_idx:
                            positions.append(i)
                        char_count += 1
                    if char_count > end_idx:
                        break
                if positions:
                    full_span = range(positions[0], positions[-1] + 1)
                    highlight_ranges.append([positions[0] + leading_columns_count + 1, positions[-1] + leading_columns_count + 1])
                    print_ranges.append([positions[0] + 1, positions[-1] + 1])
                    pattern_print_ranges[key].append((positions[0] + 1, positions[-1] + 1))
                    match_summary[key] += 1

                    for col in [pos + leading_columns_count + 1 for pos in full_span]:
                        ws.cell(row=row_idx, column=col).fill = styles['sky_blue_fill']

    highlight_time = time.time() - highlight_start
    print(f"  ⏱️  Pattern matching & highlighting time: {highlight_time:.3f} seconds")

    if not any_matches_found:
        print("\nNo patterns were found in any of the sequences.")
        highlight_ranges_freq = Counter()
    else:
        highlight_ranges_freq = Counter(map(tuple, highlight_ranges))

    return highlight_ranges, print_ranges, match_summary, pattern_print_ranges, highlight_ranges_freq, motif_hits
   
   
def apply_pattern_highlights(ws, df, pattern_label_pairs,
                             leading_columns_count, styles,
                             n_procs=None, conservation_thresh=60):
    """
    Multi-process regex matching.  
    """
    if n_procs is None or n_procs == 1:
        return _apply_pattern_highlights_ST(ws, df, pattern_label_pairs,
                                          leading_columns_count, styles)

    print(f"  🚀  Motif matching with {n_procs} processes …")
    mp_start = time.time()

    # 1. build chunks
    total_rows = len(df)
    chunk_size = max(1, total_rows // n_procs)
    chunks = []
    for i in range(n_procs):
        start = i * chunk_size
        end = None if i == n_procs - 1 else (i + 1) * chunk_size
        chunk = df.iloc[start:end].copy()
        chunks.append((chunk, pattern_label_pairs, start + 2, leading_columns_count))

    # 2. run pools
    with mp.Pool(processes=n_procs) as pool:
        results = pool.map(_match_chunk, chunks)

    # 3. merge results
    all_ranges = []
    match_summary = defaultdict(int)
    pattern_print_ranges = defaultdict(list)
    motif_hits_all = []  # NEW: (seq_index, motif_pattern, start_no_gap, end_no_gap)

    for r in results:
        all_ranges.extend(r["ranges"])
        for k, v in r["summary"].items():
            match_summary[k] += v
        for k, lst in r["pattern_ranges"].items():
            pattern_print_ranges[k].extend(lst)
        motif_hits_all.extend(r.get("motif_hits", []))

    # ⬇️ NEW: build highlight_ranges and its frequency
    highlight_ranges = []
    for (row_idx, start_cleaned, end_cleaned) in all_ranges:
        start_col = start_cleaned + leading_columns_count + 1
        end_col   = end_cleaned   + leading_columns_count + 1
        highlight_ranges.append((start_col, end_col))
    highlight_ranges_freq = Counter(highlight_ranges)
    
    # 4. OPTIMIZED: Pre-create fill object once, minimal cell access
    sky_blue = styles['sky_blue_fill']
    
    # Apply highlights - single pass
    # all_ranges items are: (excel_row_idx, start_cleaned_idx, end_cleaned_idx)
    for row_idx, start_cleaned, end_cleaned in all_ranges:
        # Convert cleaned (with-gaps) 0-based indices into excel 1-based columns
        start_excel = start_cleaned + leading_columns_count + 1
        end_excel = end_cleaned + leading_columns_count + 2  # +1 for 1-based, +1 because range end is exclusive

        for col in range(start_excel, end_excel):
            ws.cell(row=row_idx, column=col).fill = sky_blue

    mp_time = time.time() - mp_start
    print(f"  ⏱️  Parallel motif matching time: {mp_time:.3f} s")
    
    # ⬇️ return a non-empty highlight_ranges_freq
    return highlight_ranges, [], match_summary, pattern_print_ranges, highlight_ranges_freq, motif_hits_all

                            
def load_highlight_positions(positions_file):
    """Load positions to highlight from file"""
    try:
        with open(positions_file, "r") as f:
            input_positions = f.read().strip().split()
        return [int(pos) for pos in input_positions if pos.isdigit()]
    except Exception as e:
        raise PositionsFileError(f"Failed to read positions file: {str(e)}")

def apply_position_highlights(ws, highlight_cols, leading_columns_count, styles):
    """Apply position highlights to worksheet"""
    pos_start = time.time()
    
    for pos in highlight_cols:
        actual_col = get_column_letter(pos + leading_columns_count)
        for row in range(2, ws.max_row + 1):
            ws[actual_col + str(row)].fill = styles['green_fill']
    
    pos_time = time.time() - pos_start
    print(f"  ⏱️  Position highlighting time: {pos_time:.3f} seconds")

def create_summary_sheet(wb, match_summary, pattern_print_ranges, total_rows, styles, pattern_label_pairs):
    """Create match summary sheet with positional dispersion"""
    summary_start = time.time()
    
    summary_sheet = wb.create_sheet("Match Summary")
    summary_headers = ["Regex Pattern", "Matches", "Matched Subsequence Positions", "Positional Dispersion (σ)"]
    summary_sheet.append(summary_headers)

    # Apply bold style to header cells
    for cell in summary_sheet[1]:
        cell.font = styles['bold_font']

    # Fill match summary sheet
    for pattern, label in pattern_label_pairs:
        key = label or pattern
        count = match_summary[key]
        
        # Build position string for this pattern only
        ranges = pattern_print_ranges.get(key, [])
        freq_counter = Counter(ranges)
        matched_positions_str = "; ".join([f"{start}-{end} ({freq})" for (start, end), freq in sorted(freq_counter.items())])

        # Calculate positional dispersion (sigma)
        # Use start positions of all occurrences
        start_positions = [start for start, end in ranges]
        
        if len(start_positions) > 0:
            # Calculate mean position
            mean_pos = np.mean(start_positions)
            # Calculate population standard deviation (ddof=0)
            sigma = np.std(start_positions, ddof=0)
            sigma_str = f"{sigma:.2f}"
        else:
            sigma_str = "N/A"

        summary_sheet.append([
            pattern,
            count,
            matched_positions_str,
            sigma_str
        ])

    # Adjust column widths - optimized
    for col_idx, column_cells in enumerate(summary_sheet.columns, 1):
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        adjusted_width = min(max_length + 2, 50)
        col_letter = get_column_letter(col_idx)
        summary_sheet.column_dimensions[col_letter].width = adjusted_width

    # Apply alignment in bulk
    apply_bulk_alignment(summary_sheet, styles['left_alignment'], min_row=1)
    
    summary_time = time.time() - summary_start
    print(f"  ⏱️  Summary sheet creation time: {summary_time:.3f} seconds")

def run_domain_scanning(df, pfam_db_path, evalue_threshold=0.001, seq_id_offset=0):
    """Run HMMER domain scanning and return results"""
    domain_start = time.time()
    
    # Use the cleaned/processed sequences from 'protein sequences_match'
    sequences = {
        str(i + 1 + seq_id_offset): seq
        for i, seq in enumerate(df['protein sequences_match'].dropna().tolist())
    }

    fasta_path = None
    domtblout_path = None
    
    try:
        # Write sequences to a temporary FASTA file
        with tempfile.NamedTemporaryFile(mode="w+", suffix=".fasta", delete=False) as fasta_file:
            for seq_id, seq in sequences.items():
                fasta_file.write(f">{seq_id}\n{seq}\n")
            fasta_path = fasta_file.name

        # Prepare a temporary file to store domtblout results
        with tempfile.NamedTemporaryFile(mode="w+", delete=False) as domtbl_file:
            domtblout_path = domtbl_file.name

        # Run hmmscan with domtblout (no --cpu flag, let HMMER use available cores)
        cmd = [
            "hmmscan",
            "--domtblout", domtblout_path,
            pfam_db_path,
            fasta_path
        ]
        
        process = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        
        if process.returncode != 0:
            raise HMMERError(f"hmmscan failed with return code {process.returncode}:\n{process.stderr}")
        
        # Parse domtblout and filter results
        res_lines = []
        
        with open(domtblout_path, 'r') as file:
            for line in file:
                if line.startswith("#"):
                    continue
                fields = line.strip().split()
                if len(fields) < 23:
                    continue

                query_id = fields[3]
                accession = fields[1]
                domain_name = fields[0]
                description = ' '.join(fields[22:])
                c_evalue = fields[11]
                i_evalue = fields[12]
                bit_score = float(fields[13])
                start = int(fields[17])
                end = int(fields[18])

                # Filter: only store if E-value <= threshold
                try:
                    if float(i_evalue) <= evalue_threshold:
                        res_lines.append(
                            f"Sequence:{query_id} | Start:{start} | End:{end} | ID:{accession} "
                            f"| Description:{description} | E-value:{i_evalue} | Cond.E-value:{c_evalue} "
                            f"| Bit score:{bit_score}"
                        )
                except ValueError:
                    continue
        
        domain_time = time.time() - domain_start
        print(f"  ⏱️  Domain scanning time: {domain_time:.3f} seconds")
        return res_lines
        
    finally:
        # Cleanup temporary files
        try:
            if fasta_path:
                os.remove(fasta_path)
            if domtblout_path:
                os.remove(domtblout_path)
        except OSError:
            pass


def assign_sequence_chunks(total_sequences, num_processes):
    """
    Divide sequences into chunks for parallel processing.
    Returns list of (start_idx, end_idx) tuples (inclusive).
    """
    if num_processes < 1 or total_sequences == 0:
        return [(0, total_sequences - 1)]
    
    chunk_size = total_sequences // num_processes
    chunks = []
    
    for i in range(num_processes):
        start_idx = i * chunk_size
        if i == num_processes - 1:
            # Last chunk gets remaining sequences
            end_idx = total_sequences - 1
        else:
            end_idx = start_idx + chunk_size - 1
        
        chunks.append((start_idx, end_idx))
        print(f"Chunk {i}: sequences {start_idx} to {end_idx}")
    
    return chunks


def process_sequence_chunk(df_chunk, pfam_db_path, evalue_threshold, chunk_id, output_dir, offset=0):
    """
    Process a chunk of sequences with HMMER.
    Writes results to a temporary file.
    Each process runs its own hmmscan instance.
    """
    try:
        print(f"Process {chunk_id}: Starting hmmscan with {len(df_chunk)} sequences")
        chunk_start = time.time()
        
        # Run domain scanning on this chunk (runs independent hmmscan)
        res_lines = run_domain_scanning(df_chunk, pfam_db_path, evalue_threshold, seq_id_offset=offset)
        
        # Write results to temporary file
        temp_file = os.path.join(output_dir, f'chunk_{chunk_id}_results.tmp')
        with open(temp_file, 'w') as f:
            for line in res_lines:
                f.write(line + '\n')
        
        chunk_time = time.time() - chunk_start
        print(f"Process {chunk_id}: Completed in {chunk_time:.2f}s. Found {len(res_lines)} domain hits")
        return len(res_lines)
        
    except Exception as e:
        print(f"Process {chunk_id}: Error - {str(e)}")
        return 0


def merge_chunk_results(output_dir):
    """
    Merge all chunk results from temporary files.
    Returns combined list of result lines.
    """
    print("Merging results from all chunks...")
    all_results = []
    
    # Find all temporary result files
    temp_files = sorted([f for f in os.listdir(output_dir) if f.endswith('.tmp')])
    
    for temp_file in temp_files:
        file_path = os.path.join(output_dir, temp_file)
        try:
            with open(file_path, 'r') as f:
                lines = f.readlines()
                all_results.extend([line.strip() for line in lines if line.strip()])
            
            # Remove temporary file after reading
            os.remove(file_path)
        except Exception as e:
            print(f"Error reading {temp_file}: {e}")
    
    print(f"Merged {len(all_results)} total domain hits")
    return all_results


def run_parallel_domain_scanning(df, pfam_db_path, evalue_threshold, num_processes):
    """
    Run domain scanning in parallel using multiprocessing.
    Each process will run its own independent hmmscan instance.
    
    Args:
        df: DataFrame with sequences
        pfam_db_path: Path to Pfam database
        evalue_threshold: E-value cutoff
        num_processes: Number of parallel processes (hmmscan instances)
    
    Returns:
        Combined list of domain hits from all processes
    """
    parallel_start = time.time()
    
    total_sequences = len(df)
    print(f"\n{'='*60}")
    print(f"PARALLEL DOMAIN SCANNING")
    print(f"{'='*60}")
    print(f"Total sequences: {total_sequences}")
    print(f"Number of parallel hmmscan instances: {num_processes}")
    
    # Create temporary directory for chunk results
    temp_dir = tempfile.mkdtemp(prefix='domain_scan_')
    print(f"Temporary directory: {temp_dir}")
    
    # Divide sequences into chunks
    chunks = assign_sequence_chunks(total_sequences, num_processes)
    
    # Create processes for each chunk
    processes = []
    for i, (start_idx, end_idx) in enumerate(chunks):
        # Extract chunk dataframe
        df_chunk = df.iloc[start_idx:end_idx + 1].copy()
        # Reset index for the chunk so sequence IDs match within chunk
        #df_chunk.reset_index(drop=True, inplace=True)
        
        # Adjust sequence IDs to match original positions
        # We'll need to track the offset
        offset = start_idx
        
        # Create process - each will run its own hmmscan
        p = multiprocessing.Process(
            target=process_sequence_chunk,
            args=(df_chunk, pfam_db_path, evalue_threshold, i, temp_dir, start_idx)
        )
        processes.append(p)
        p.start()
        print(f"Started hmmscan process {i} (PID: {p.pid}) for sequences {start_idx}-{end_idx}")
    
    # Wait for all processes to complete
    print(f"\nWaiting for all {num_processes} hmmscan processes to complete...")
    for i, p in enumerate(processes):
        p.join()
        print(f"✓ hmmscan process {i} (PID: {p.pid}) completed")
    
    # Merge results from all chunks
    all_results = merge_chunk_results(temp_dir)
    
    # Cleanup temporary directory
    try:
        os.rmdir(temp_dir)
    except:
        pass
    
    parallel_time = time.time() - parallel_start
    print(f"\n{'='*60}")
    print(f"Parallel scanning completed in {parallel_time:.3f} seconds")
    print(f"Total speedup vs serial: ~{num_processes}x (theoretical)")
    print(f"{'='*60}\n")
    
    return all_results

def apply_domain_highlights(wb, ws, df, res_lines, leading_columns_count, styles, df_final):
    """OPTIMIZED: Apply domain highlights by creating new sheet from scratch"""
    domain_highlight_start = time.time()
    
    domain_highlights = []
    
    for line in res_lines:
        try:
            parts = line.split('|')
            seq_num = parts[0].split(':')[1].strip()
            start_pos = int(parts[1].split(':')[1].strip())
            end_pos = int(parts[2].split(':')[1].strip())
            accession_id = parts[3].split(':')[1].strip()
            domain_name = parts[4].split(':')[1].strip()
            domain_highlights.append((seq_num, start_pos, end_pos, domain_name, accession_id))
        except (IndexError, ValueError):
            continue

    if not domain_highlights:
        return False

    # OPTIMIZATION: Create new sheet directly instead of copying
    create_start = time.time()
    domain_ws = wb.create_sheet("Domain_Highlights")
    
    # Write data directly from dataframe
    for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 1):
        domain_ws.append(row)
    
    print(f"    └─ Sheet creation time: {time.time() - create_start:.3f}s")

    # Build position cache once for ALL sequences
    cache_start = time.time()
    position_cache = {}
    
    for idx in range(len(df)):
        seq_num = str(idx + 1)
        cleaned_seq = df.iloc[idx]['protein sequences_cleaned']
        
        # Build position mapping
        pos_map = {}
        char_idx = 0
        for i, char in enumerate(cleaned_seq):
            if char not in {" ", "-"}:
                char_idx += 1
                pos_map[char_idx] = i
        position_cache[seq_num] = (cleaned_seq, pos_map)
    
    print(f"    └─ Position cache build time: {time.time() - cache_start:.3f}s")

    # Group all domains by sequence number (row)
    highlight_start = time.time()
    domains_by_row = defaultdict(list)
    for seq_num, start_pos, end_pos, domain_name, accession_id in domain_highlights:
        if int(seq_num) < 1 or int(seq_num) > len(df):
            continue
        domains_by_row[seq_num].append((start_pos, end_pos, domain_name, accession_id))
    
    # Pre-create fill and comment objects
    orange_fill = styles['orange_fill']
    
    # Process each row only once
    for seq_num, domains in domains_by_row.items():
        row_idx = int(seq_num) + 1
        cleaned_seq, pos_map = position_cache[seq_num]
        
        # Get all cells for this row at once
        row_cells = list(domain_ws[row_idx])
        
        # Process all domains for this row
        for start_pos, end_pos, domain_name, accession_id in domains:
            domain_start_idx = pos_map.get(start_pos)
            domain_end_idx = pos_map.get(end_pos)
            
            if domain_start_idx is None or domain_end_idx is None:
                continue
            
            # Highlight cells
            for i in range(domain_start_idx, domain_end_idx + 1):
                actual_col_idx = i + leading_columns_count
                if actual_col_idx < len(row_cells):
                    cell = row_cells[actual_col_idx]
                    cell.fill = orange_fill
                    
                    # Add comment only to first cell of domain
                    if i == domain_start_idx:
                        aligned_start = domain_start_idx + 1
                        aligned_end = domain_end_idx + 1
                        comment_text = (
                            f"Domain: {domain_name}\n"
                            f"Accession: {accession_id}\n"
                            f"Start (with gaps): {aligned_start}\n"
                            f"End (with gaps): {aligned_end}"
                        )
                        cell.comment = Comment(comment_text, "DomainInfo")
    
    print(f"    └─ Domain highlighting time: {time.time() - highlight_start:.3f}s")
    
    # Apply bold and center alignment to header row
    for cell in domain_ws[1]:
        cell.font = styles['bold_font']
        cell.alignment = styles['center_alignment']

    # OPTIMIZATION: Apply styling once at the end
    style_start = time.time()
    adjust_column_widths(domain_ws)
    apply_bulk_alignment(domain_ws, styles['center_alignment'], min_row=2)
    print(f"    └─ Styling time: {time.time() - style_start:.3f}s")
    
    domain_highlight_time = time.time() - domain_highlight_start
    print(f"  ⏱️  Total domain highlighting time: {domain_highlight_time:.3f} seconds")
    print("\nFinished creating 'Domain_Highlights' sheet.")
    
    # Create 'Domain Summary' sheet
    create_domain_summary_sheet(wb, res_lines, styles)
    
    return True

def create_domain_summary_sheet(wb, res_lines, styles):
    """Create domain summary sheet with full domain information"""
    domain_summary_start = time.time()
    
    domain_summary_sheet = wb.create_sheet("Domain Summary")

    headers = [
        "Sequence", "Start(without '-')", "End(without '-')",
        "Domain Name", "Accession ID",
        "E-value", "Cond. E-value", "Bit Score"
    ]
    domain_summary_sheet.append(headers)

    # Apply bold style to header cells
    for cell in domain_summary_sheet[1]:
        cell.font = styles['bold_font']

    for line in res_lines:
        try:
            parts = line.split('|')

            query_id = parts[0].split(':')[1].strip()
            start = int(parts[1].split(':')[1].strip())
            end = int(parts[2].split(':')[1].strip())
            accession = parts[3].split(':')[1].strip()
            description = parts[4].split(':')[1].strip()
            i_evalue = float(parts[5].split(':')[1].strip())
            c_evalue = float(parts[6].split(':')[1].strip())
            bit_score = float(parts[7].split(':')[1].strip())

            domain_summary_sheet.append([
                query_id, start, end,
                description, accession,
                i_evalue, c_evalue, bit_score
            ])
        except (IndexError, ValueError):
            continue
    
    # Adjust column widths - optimized
    for col_idx, column_cells in enumerate(domain_summary_sheet.columns, 1):
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        adjusted_width = min(max_length + 2, 50)
        col_letter = get_column_letter(col_idx)
        domain_summary_sheet.column_dimensions[col_letter].width = adjusted_width

    # Apply alignment in bulk
    apply_bulk_alignment(domain_summary_sheet, styles['left_alignment'], min_row=1)
    
    domain_summary_time = time.time() - domain_summary_start
    print(f"  ⏱️  Domain summary sheet creation time: {domain_summary_time:.3f} seconds")
    
def _parse_domains_by_sequence(res_lines):
    """Parse HMMER result lines into {seq_index(int): [(accession,str), start(int), end(int)]} (gapless coords)."""
    domains_by_seq = defaultdict(list)
    for line in res_lines or []:
        try:
            parts = line.split('|')
            seq_num = int(parts[0].split(':')[1].strip())
            start_pos = int(parts[1].split(':')[1].strip())
            end_pos = int(parts[2].split(':')[1].strip())
            accession_id = parts[3].split(':')[1].strip()
            # Store gapless coordinates
            domains_by_seq[seq_num].append((accession_id, start_pos, end_pos))
        except Exception:
            continue
    # Sort domains by start for a tiny speed win
    for k in domains_by_seq:
        domains_by_seq[k].sort(key=lambda x: x[1])
    return domains_by_seq


def _compute_mdcs_summary(motif_hits, domains_by_seq):
    """Returns MDCS rows + DCM counters for the second table + motif statistics for third table"""
    # Group motifs by sequence
    motifs_by_seq = defaultdict(lambda: defaultdict(list))
    
    # FIXED: 4-element tuple unpacking (no extra "_")
    for seq_idx, motif_pat, m_start, m_end in motif_hits:
        motifs_by_seq[int(seq_idx)][motif_pat].append((int(m_start), int(m_end)))

    rows = []
    
    # DCM counters: domain -> motif -> full/partial/total count
    full_count = defaultdict(lambda: defaultdict(int))
    partial_count = defaultdict(lambda: defaultdict(int))
    total_domain_count = defaultdict(lambda: defaultdict(int))  # NEW: total overlaps per domain
    
    # Motif statistics: motif -> list of MDCS values and counts
    motif_mdcs_values = defaultdict(list)
    motif_outside_count = defaultdict(int)  # NEW: count of outside domain occurrences
    motif_touching_count = defaultdict(int)  # NEW: count of domain-touching occurrences

    for seq_idx in sorted(motifs_by_seq.keys()):
        doms = domains_by_seq.get(seq_idx, [])
        has_domains = len(doms) > 0

        for motif_pat in sorted(motifs_by_seq[seq_idx].keys()):
            occs = motifs_by_seq[seq_idx][motif_pat]

            best_by_domain = {}
            has_outside_occurrence = False
            
            if has_domains:
                for acc, d_start, d_end in doms:
                    best = 0.0
                    for m_start, m_end in occs:
                        motif_len = m_end - m_start + 1
                        if motif_len <= 0:
                            continue
                        ov = max(0, min(m_end, d_end) - max(m_start, d_start) + 1)
                        mdcs = ov / motif_len
                        if mdcs > best:
                            best = mdcs
                    if best > 0:
                        best_by_domain[acc] = max(best_by_domain.get(acc, 0.0), best)
                
                # Check if any occurrence is completely outside all domains
                for m_start, m_end in occs:
                    is_outside = True
                    for acc, d_start, d_end in doms:
                        ov = max(0, min(m_end, d_end) - max(m_start, d_start) + 1)
                        if ov > 0:
                            is_outside = False
                            break
                    if is_outside:
                        has_outside_occurrence = True
                        break

            # Build MDCS cell and interpretation
            if best_by_domain:
                items = sorted(best_by_domain.items(), key=lambda x: (-x[1], x[0]))
                mdcs_cell = " , ".join([f"{val:.2f}--{acc}" for acc, val in items])
                max_val = max(best_by_domain.values())
                interpretation = "Fully Embedded" if abs(max_val - 1.0) < 1e-9 else "Partial Overlap"
                # Count one occurrence per unique domain (with best MDCS for that domain)
                for domain_acc, mdcs_val in best_by_domain.items():
                    motif_mdcs_values[motif_pat].append(mdcs_val)
                    motif_touching_count[motif_pat] += 1  # Count domain-touching occurrences
            else:
                mdcs_cell = "0"
                interpretation = "No Domains" if not has_domains else "Outside Domains"
                # Count as one occurrence with MDCS=0 for this sequence
                if has_domains:
                    motif_mdcs_values[motif_pat].append(0.0)
                    motif_outside_count[motif_pat] += 1  # Count outside domain occurrences
            
            # Additional occurrence if motif is also found outside all domains
            # (even when it overlaps with some domains in the same sequence)
            if has_domains and has_outside_occurrence and best_by_domain:
                motif_mdcs_values[motif_pat].append(0.0)
                motif_outside_count[motif_pat] += 1

            rows.append([seq_idx, motif_pat, mdcs_cell, interpretation])

            # Efficient DCM counting
            for acc, mdcs_val in best_by_domain.items():
                total_domain_count[acc][motif_pat] += 1  # NEW: Count total overlaps
                if abs(mdcs_val - 1.0) < 1e-9:
                    full_count[acc][motif_pat] += 1
                else:
                    partial_count[acc][motif_pat] += 1

    return rows, full_count, partial_count, total_domain_count, motif_mdcs_values, motif_outside_count, motif_touching_count


def create_mdcs_summary_sheet(wb, motif_hits, res_lines, styles):
    """MDCS Summary sheet with Motif Embedding in Domain table and MDCS Statistics table"""
    domains_by_seq = _parse_domains_by_sequence(res_lines)
    rows, full_count, partial_count, total_domain_count, motif_mdcs_values, motif_outside_count, motif_touching_count = _compute_mdcs_summary(motif_hits, domains_by_seq)

    mdcs_ws = wb.create_sheet("MDCS Summary")

    # First table: MDCS
    headers = ["Sequence", "Motif pattern", "MDCS", "Interpretation"]
    mdcs_ws.append(headers)
    for cell in mdcs_ws[1]:
        cell.font = styles['bold_font']

    for r in rows:
        mdcs_ws.append(r)

    # 3 blank rows gap
    mdcs_ws.append([])
    mdcs_ws.append([])
    mdcs_ws.append([])

    # Second table: Motif Embedding in Domain (with Count column)
    dcm_headers = ["Domain", "Motif Pattern", "Motif Embedding in Domain", "Touching This Domain"]
    mdcs_ws.append(dcm_headers)
    for cell in mdcs_ws[mdcs_ws.max_row]:
        cell.font = styles['bold_font']

    # Fill Motif Embedding table
    all_domains = sorted(set(full_count.keys()) | set(partial_count.keys()))
    for dom in all_domains:
        all_motifs = sorted(set(full_count[dom].keys()) | set(partial_count[dom].keys()))
        for mot in all_motifs:
            full = full_count[dom][mot]
            partial = partial_count[dom][mot]
            total = full + partial
            if total == 0:
                continue
            pct_full = (full / total) * 100
            pct_partial = (partial / total) * 100
            dcm_text = f"full={pct_full:.0f}% times , partial={pct_partial:.0f}% times"
            overlap_count = f"{total_domain_count[dom][mot]} times"  # Total times this motif overlaps this domain
            mdcs_ws.append([dom, mot, dcm_text, overlap_count])

    # 3 blank rows gap
    mdcs_ws.append([])
    mdcs_ws.append([])
    mdcs_ws.append([])

    # Third table: MDCS Statistics per Motif (with Outside and Touching counts)
    import numpy as np
    stats_headers = ["", "Motif Pattern", "Mean MDCS", "Median MDCS", "Occurrences", "Outside Domains", "Touching Domains"]
    mdcs_ws.append(stats_headers)
    for cell in mdcs_ws[mdcs_ws.max_row]:
        cell.font = styles['bold_font']

    # Calculate statistics for each motif
    for motif_pat in sorted(motif_mdcs_values.keys()):
        values = motif_mdcs_values[motif_pat]
        if values:
            mean_mdcs = np.mean(values)
            median_mdcs = np.median(values)
            count = len(values)
            outside = motif_outside_count.get(motif_pat, 0)
            touching = motif_touching_count.get(motif_pat, 0)
            
            mdcs_ws.append([
                "",  # First column blank
                motif_pat,
                f"{mean_mdcs:.3f}",
                f"{median_mdcs:.3f}",
                count,
                outside,
                touching
            ])

    # Final styling
    adjust_column_widths(mdcs_ws)
    apply_bulk_alignment(mdcs_ws, styles['center_alignment'], min_row=1)
    return True



print("\nFinished creating 'Domain Summary' sheet.")


class ProteinAnalysisApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Protein Sequence Analysis Tool")
        self.root.geometry("700x600")
        self.root.minsize(650, 600)
        
        self.style = ttk.Style()
        self.style.configure('TLabelframe.Label', font=('Arial', 12, 'bold'))
        self.style.configure('TLabel', font=('Segoe UI', 11))
        self.style.configure('TButton', font=('Helvetica', 11, 'bold'), borderwidth=3, relief='raised', focuscolor='#2196F3')
        self.style.map('TButton', 
                      background=[('active', '#e3f2fd'), ('pressed', "#bbdefb")],
                      bordercolor=[('active', '#2196F3'), ('pressed', '#1976D2'), ('!active', '#4CAF50')])
        self.style.configure('TCheckbutton', font=('Segoe UI', 11))

        # Variables
        self.motif_search_var = tk.BooleanVar()
        self.position_highlight_var = tk.BooleanVar()
        self.domain_scanning_var = tk.BooleanVar()
        self.motif_file_path = tk.StringVar()
        self.regex_patterns = tk.StringVar()
        self.conservation_threshold = tk.StringVar(value="60")
        self.positions_to_highlight = tk.StringVar()
        self.pfam_db_path = tk.StringVar()
        self.evalue_threshold = tk.StringVar(value="0.001")
        self.motif_parallel_processes = tk.StringVar(value=str(max(1, multiprocessing.cpu_count() // 2)))
        self.num_processes = tk.StringVar(value=str(max(1, multiprocessing.cpu_count() // 2)))


        self.progress_window = None
        
        # Create scrollable main frame
        self.create_scrollable_frame()
        
        # Create GUI elements
        self.create_widgets()
        
        # Initialize proceed button state
        self.proceed_button['state'] = tk.DISABLED
        
        # Set up trace calls after all methods are defined
        self.setup_traces()
    
    def create_scrollable_frame(self):
        """Create a scrollable main frame"""
        # Create main canvas and scrollbar
        self.main_canvas = tk.Canvas(self.root, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.main_canvas.yview)
        self.scrollable_frame = ttk.Frame(self.main_canvas, padding="5")
        
        # Configure scrolling
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
        )
        
        # Create window and store the window ID
        self.canvas_window = self.main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.main_canvas.configure(yscrollcommand=self.scrollbar.set)
        
        # Pack canvas and scrollbar
        self.main_canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar.pack(side="right", fill="y")
        
        # Bind canvas resize to update frame width
        def _on_canvas_configure(event):
            canvas_width = event.width
            self.main_canvas.itemconfig(self.canvas_window, width=canvas_width)
        
        self.main_canvas.bind('<Configure>', _on_canvas_configure)
        
        # Bind mouse wheel to canvas only when scrolling is needed
        def _on_mousewheel(event):
            if self.main_canvas.bbox("all")[3] > self.main_canvas.winfo_height():
                self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def _bind_to_mousewheel(event):
            self.main_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        def _unbind_from_mousewheel(event):
            self.main_canvas.unbind_all("<MouseWheel>")
        
        self.main_canvas.bind('<Enter>', _bind_to_mousewheel)
        self.main_canvas.bind('<Leave>', _unbind_from_mousewheel)
    
    def setup_traces(self):
        """Set up variable traces after all methods are defined"""
        self.motif_file_path.trace_add('write', lambda *args: self.check_inputs())
        self.positions_to_highlight.trace_add('write', lambda *args: self.check_inputs())
        self.pfam_db_path.trace_add('write', lambda *args: self.check_inputs())
        self.regex_text.bind('<KeyRelease>', lambda *args: self.check_inputs())
    
    def create_widgets(self):
        """Create all GUI widgets"""
        # 1. File Selection
        file_frame = ttk.LabelFrame(self.scrollable_frame, text="1. Select Input File", padding="9")
        file_frame.pack(fill=tk.X, pady=(0,10))
        
        ttk.Label(file_frame, text="Input File (Excel or FASTA):").pack(anchor="w")
        file_entry_frame = ttk.Frame(file_frame)
        file_entry_frame.pack(fill=tk.X, pady=5)

        ttk.Entry(file_entry_frame, textvariable=self.motif_file_path, font=('Arial', 10)).pack(side=tk.LEFT, fill=tk.X, expand=True)
        browse_btn1 = ttk.Button(file_entry_frame, text="Browse",
                      command=lambda: self.browse_file(self.motif_file_path))
        browse_btn1.pack(side=tk.RIGHT)
        
        # 2. Analysis Options
        options_frame = ttk.LabelFrame(self.scrollable_frame, text="2. Select Analysis Options", padding="9")
        options_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Checkbutton(options_frame, text="Motif Searching", variable=self.motif_search_var,
                       command=self.toggle_motif_options).pack(anchor="w")
        ttk.Checkbutton(options_frame, text="Position Highlighting", variable=self.position_highlight_var,
                       command=self.toggle_position_options).pack(anchor="w")
        ttk.Checkbutton(options_frame, text="Domain Scanning (HMMER)", variable=self.domain_scanning_var,
                       command=self.toggle_domain_options).pack(anchor="w", pady=(0, 5))

        # 3. Motif Options
        self.motif_frame = ttk.LabelFrame(self.scrollable_frame, text="Motif Search Options", padding="9")
        
        ttk.Label(self.motif_frame, text="Regex Patterns (one per line):").pack(anchor="w",pady=(0,5))
        self.regex_text = tk.Text(self.motif_frame, height=6, width=60, font=('Consolas', 11), wrap=tk.NONE)
        self.regex_text.pack(fill=tk.X, expand=True, pady=5)
        
         # Threshold setting
        threshold_frame = ttk.Frame(self.motif_frame)
        threshold_frame.pack(fill=tk.X,pady=(5,0))
        ttk.Label(threshold_frame, text="Conservation Threshold (%):")\
            .pack(side=tk.LEFT)
        ttk.Entry(threshold_frame, textvariable=self.conservation_threshold, width=4, font=('Arial', 10))\
            .pack(side=tk.LEFT, padx=8)
            
        # ---- parallel processes for motif search -----------------
        motif_parallel_frame = ttk.Frame(self.motif_frame)
        motif_parallel_frame.pack(fill=tk.X, pady=(5,0))
        ttk.Label(motif_parallel_frame, text="Parallel processes for motif search:").pack(side=tk.LEFT)
        ttk.Entry(motif_parallel_frame, textvariable=self.motif_parallel_processes,
                  width=5, font=('Arial',10)).pack(side=tk.LEFT, padx=(5,0))
        ttk.Label(motif_parallel_frame,
                  text=f"(CPUs: {multiprocessing.cpu_count()})",
                  foreground='gray', font=('Segoe UI',9)).pack(side=tk.LEFT, padx=(5,0))
            
        # 4. Position Options
        self.position_frame = ttk.LabelFrame(self.scrollable_frame, text="Position Highlight Options", padding="9")
        ttk.Label(self.position_frame, text="Positions to Highlight (space separated):").pack(anchor="w")
        ttk.Entry(self.position_frame, textvariable=self.positions_to_highlight, font=('Arial', 10))\
            .pack(fill=tk.X,pady=6)
        
        # 5. Domain Scanning Options
        self.domain_frame = ttk.LabelFrame(self.scrollable_frame, text="Domain Scanning Options", padding="9")
        
        ttk.Label(self.domain_frame, text="Pfam Database Path:").pack(anchor="w")
        pfam_entry_frame = ttk.Frame(self.domain_frame)
        pfam_entry_frame.pack(fill=tk.X, pady=5)
        
        ttk.Entry(pfam_entry_frame, textvariable=self.pfam_db_path, font=('Arial', 10)).pack(side=tk.LEFT, fill=tk.X, expand=True)
        browse_btn2 = ttk.Button(pfam_entry_frame, text="Browse", 
                      command=lambda: self.browse_file(self.pfam_db_path, [("HMM files", "*.hmm"), ("All files", "*.*")]))
        browse_btn2.pack(side=tk.RIGHT)
        
        # Domain scanning parameters
        domain_params_frame = ttk.Frame(self.domain_frame)
        domain_params_frame.pack(fill=tk.X, pady=(5,0))
        
        ttk.Label(domain_params_frame, text="E-value Threshold:").pack(side=tk.LEFT)
        ttk.Entry(domain_params_frame, textvariable=self.evalue_threshold, width=8, font=('Arial', 10))\
            .pack(side=tk.LEFT, padx=(5,20))
            
        ttk.Label(domain_params_frame, text="Parallel Processes:").pack(side=tk.LEFT)
        ttk.Entry(domain_params_frame, textvariable=self.num_processes, width=5, font=('Arial', 10))\
            .pack(side=tk.LEFT, padx=(5,0))
            
        # Info label for recommended processes
        cpu_count = multiprocessing.cpu_count()
        info_label = ttk.Label(domain_params_frame, 
                               text=f"(CPUs: {cpu_count})",
                               font=('Segoe UI', 9), 
                               foreground='gray')
        info_label.pack(side=tk.LEFT, padx=(5,0))
            
        # Create a container frame for bottom elements
        bottom_frame = ttk.Frame(self.scrollable_frame, padding=10)
        bottom_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(10, 0))
        
        # Add progress bar (initially hidden)
        self.progress = ttk.Progressbar(bottom_frame, mode='indeterminate')
        
        # Add runtime label
        self.runtime_label = ttk.Label(bottom_frame, text="", font=('Segoe UI', 9))
        
        # Proceed button now in bottom right
        button_frame = ttk.Frame(bottom_frame)
        button_frame.pack(side=tk.RIGHT)
        
        self.proceed_button = ttk.Button(button_frame, text="Proceed", command=self.run_analysis)
        self.proceed_button.pack()
        
        # Initially hide optional frames
        self.motif_frame.pack_forget()
        self.position_frame.pack_forget()
        self.domain_frame.pack_forget()
        
        # Update canvas scroll region after initial widget creation
        self.scrollable_frame.update_idletasks()
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
    
    def browse_file(self, target_variable, filetypes=None):
        """Open file dialog and set target variable"""
        if filetypes is None:
            filetypes = [("Supported files", "*.xlsx *.xls *.fasta *.fa *.faa"), 
                        ("Excel files", "*.xlsx *.xls"),
                        ("FASTA files", "*.fasta *.fa *.faa"),
                        ("All files", "*.*")]
        
        file_path = filedialog.askopenfilename(filetypes=filetypes)
        if file_path:
            target_variable.set(file_path)
    
    def toggle_motif_options(self):
        """Toggle visibility of motif options frame"""
        if self.motif_search_var.get():
            options_frame = None
            for child in self.scrollable_frame.winfo_children():
                if isinstance(child, ttk.LabelFrame) and "Select Analysis Options" in str(child.cget("text")):
                    options_frame = child
                    break
            
            if options_frame:
                self.motif_frame.pack(fill=tk.X, pady=(0, 10), after=options_frame)
        else:
            self.motif_frame.pack_forget()
        
        self.update_scroll_region()
        self.check_inputs()
    
    def toggle_position_options(self):
        """Toggle visibility of position options frame"""
        if self.position_highlight_var.get():
            after_widget = None
            for child in self.scrollable_frame.winfo_children():
                if isinstance(child, ttk.LabelFrame) and "Select Analysis Options" in str(child.cget("text")):
                    after_widget = child
                    break
            
            if self.motif_search_var.get():
                after_widget = self.motif_frame
            
            if after_widget:
                self.position_frame.pack(fill=tk.X, pady=(0, 10), after=after_widget)
        else:
            self.position_frame.pack_forget()
        
        self.update_scroll_region()
        self.check_inputs()
    
    def toggle_domain_options(self):
        """Toggle visibility of domain options frame"""
        if self.domain_scanning_var.get():
            after_widget = None
            for child in self.scrollable_frame.winfo_children():
                if isinstance(child, ttk.LabelFrame) and "Select Analysis Options" in str(child.cget("text")):
                    after_widget = child
                    break
            
            if self.position_highlight_var.get():
                after_widget = self.position_frame
            elif self.motif_search_var.get():
                after_widget = self.motif_frame
            
            if after_widget:
                self.domain_frame.pack(fill=tk.X, pady=(0, 10), after=after_widget)
        else:
            self.domain_frame.pack_forget()
        
        self.update_scroll_region()
        self.check_inputs()
    
    def update_scroll_region(self):
        """Update canvas scroll region after widget changes"""
        self.scrollable_frame.update_idletasks()
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
    
    def check_inputs(self, *args):
        """Enable proceed button only when all required inputs are provided"""
        enabled = True
        
        # Check if at least one option is selected
        if not (self.motif_search_var.get() or self.position_highlight_var.get() or self.domain_scanning_var.get()):
            enabled = False
        
        # Check file is selected
        if not self.motif_file_path.get():
            enabled = False
        
        # Check motif search requirements if selected
        if self.motif_search_var.get():
            if not self.regex_text.get("1.0", tk.END).strip():
                enabled = False
        
        # Check position highlight requirements if selected
        if self.position_highlight_var.get():
            if not self.positions_to_highlight.get().strip():
                enabled = False
        
        # Check domain scanning requirements if selected
        if self.domain_scanning_var.get():
            if not self.pfam_db_path.get().strip():
                enabled = False
        
        self.proceed_button['state'] = tk.NORMAL if enabled else tk.DISABLED
    
    def run_analysis(self):
        """Execute the analysis with current parameters"""
        try:
            # Disable button during processing
            self.proceed_button['state'] = tk.DISABLED
            
            # Create progress window
            self.show_progress_window()
            
            # Record start time
            start_time = time.time()
            
            # Run analysis in a separate thread to keep UI responsive
            self._thread = threading.Thread(
                target=self.execute_analysis_wrapper,
                args=(start_time,),
                daemon=True
            )
            self._thread.start()
            
        except Exception as e:
            self.cleanup_after_analysis()
            messagebox.showerror("Error", f"An error occurred:\n{str(e)}")    
            
    def execute_analysis_wrapper(self, start_time):
        """
        Background-thread entry point.
        ALL Tkinter actions are scheduled with root.after(0, …).
        All original exception types are preserved.
        """
        try:
            # 1. GUI – disable button + show progress (main thread)
            self.root.after(0, lambda: self.proceed_button.config(state='disabled'))
            
            # 2. Build parameters exactly like the old code
            regex_input = None
            if self.motif_search_var.get():
                regex_text = self.regex_text.get("1.0", tk.END).strip()
                if not regex_text:
                    raise ValueError("Please enter at least one regex pattern for motif searching")
                regex_input = [line.strip() for line in regex_text.splitlines() if line.strip()]
                try:
                    conservation_threshold = float(self.conservation_threshold.get())
                    if not 0 <= conservation_threshold <= 100:
                        raise ValueError
                except ValueError:
                    raise ValueError("Conservation threshold must be a number between 0 and 100")

            positions_to_highlight = None
            if self.position_highlight_var.get():
                pos_text = self.positions_to_highlight.get().strip()
                if not pos_text:
                    raise ValueError("Please enter at least one position for highlighting")
                try:
                    positions_to_highlight = [int(pos) for pos in pos_text.split()]
                except ValueError:
                    raise ValueError("Positions must be space-separated numbers")

            domain_params = None
            if self.domain_scanning_var.get():
                if not self.pfam_db_path.get().strip():
                    raise ValueError("Please select a Pfam database file")
                try:
                    evalue = float(self.evalue_threshold.get())
                    if evalue <= 0:
                        raise ValueError
                except ValueError:
                    raise ValueError("E-value threshold must be a positive number")
                
                try:
                    num_proc = int(self.num_processes.get())
                    if num_proc <= 0:
                        raise ValueError
                except ValueError:
                    raise ValueError("Number of processes must be a positive integer")
            
                # HMMER presence check
                try:
                    subprocess.run(["hmmscan", "-h"], stdout=subprocess.PIPE, stderr=subprocess.PIPE, check=True)
                except (subprocess.CalledProcessError, FileNotFoundError):
                    raise HMMERError("hmmscan not found. Please ensure HMMER is installed and in your PATH.")
                if not os.path.exists(self.pfam_db_path.get()):
                    raise FileNotFoundError("Pfam database file not found. Please check the path.")
                domain_params = {
                'pfam_db_path': self.pfam_db_path.get(),
                'evalue_threshold': evalue,
                'num_processes': num_proc
                }

            # 3. Heavy work – background thread is OK
            output_file = self.execute_analysis(
                input_excel_file=self.motif_file_path.get(),
                regex_input=regex_input,
                positions_to_highlight=positions_to_highlight,
                conservation_threshold=conservation_threshold if self.motif_search_var.get() else 60,
                domain_params=domain_params
            )

            # 4. Success – show completion on main thread
            runtime = time.time() - start_time
            self.root.after(0, lambda: self.show_completion(output_file, runtime))

        # 5. Original exception hierarchy – still running in background thread
        except HeaderNotFoundError as e:
            self.root.after(0, lambda msg=str(e): messagebox.showerror("File Format Error", msg))
        except FileReadError as e:
            self.root.after(0, lambda msg=str(e): messagebox.showerror("File Read Error", msg))
        except HMMERError as e:
            self.root.after(0, lambda msg=str(e): messagebox.showerror("HMMER Error", msg))
        except ValueError as e:
            self.root.after(0, lambda msg=str(e): messagebox.showerror("Validation Error", msg))
        except Exception as e:
            self.root.after(0, lambda msg=str(e): messagebox.showerror("Unexpected Error", msg))

        # 6. Always re-enable button + close progress (main thread)
        finally:
            self.root.after(0, lambda: self.proceed_button.config(state='normal'))
            self.root.after(0, self.close_progress_window)   # BLOCK 3 helper

    def close_progress_window(self):
            """Destroy progress window – safe to call from main thread."""
            try:
                if self.progress_window and self.progress_window.winfo_exists():
                    self.progress_window.destroy()
            except tk.TclError:
                pass   # already gone
            self.progress_window = None

    def show_progress_window(self):
            """Create a dedicated progress window"""
            self.progress_window = tk.Toplevel(self.root)
            self.progress_window.title("Processing Analysis")
            self.progress_window.geometry("400x150")
            self.progress_window.resizable(False, False)
        
            # Center the window
            self.progress_window.transient(self.root)  # stay on-top
            self.progress_window.lift()
        
            # Progress label
            ttk.Label(self.progress_window, 
                 text="Processing your analysis...", 
                 font=('Segoe UI', 12)).pack(pady=(20, 10))
        
            # Green progress bar
            self.style.configure('Green.Horizontal.TProgressbar', 
                           background='#28a745',
                           thickness=20)
            self.progress = ttk.Progressbar(
            self.progress_window,
            style='Green.Horizontal.TProgressbar',
            mode='indeterminate',
            length=300
            )
            self.progress.pack(pady=10)
            self.progress.start(10)  # Start with faster animation
        
    def open_file(self, path, parent_window=None):
        """Cross-platform safe file opener with error dialog"""
        try:
            system = platform.system()
            if system == "Windows":
                os.startfile(path)
            elif system == "Darwin":  # macOS
                subprocess.run(["open", path], check=True)
            else:  # Linux or WSL
                if "microsoft" in platform.release().lower():
                    # Convert Linux path to Windows path
                    win_path = subprocess.check_output(["wslpath", "-w", path]).decode().strip()
                    subprocess.run(["explorer.exe", win_path], check=False)
                else:
                    subprocess.run(["xdg-open", path], check=True)  # native Linux
        except Exception as e:
            # show Tkinter error popup if parent_window exists
            if parent_window:
                messagebox.showerror(
                "Open Error",
                f"Cannot open file:\n{str(e)}",
                parent=parent_window
                )
            else:
                print(f"Error opening file: {e}")
 
    
    def show_completion(self, output_file, runtime):
        """Professional success dialog with large text"""
        win = tk.Toplevel(self.root)
        win.geometry("500x250")
        win.resizable(True, True)
    
        # Make this window modal (blocks interaction with main window)
        win.grab_set()
    
        # Center the window
        self._center_window(win)
    
        self.style.configure('Success.Header.TLabel', 
                       font=('Segoe UI', 12, 'bold'),
                       foreground='#2e8b57')  # Sea green color
        self.style.configure('Success.TLabel', 
                       font=('Segoe UI', 11),
                       wraplength=450)
    
        # Create frame for better content organization
        content_frame = ttk.Frame(win)
        content_frame.pack(fill='both',expand=True, padx=20, pady=15)
    
        # Success message with green text
        ttk.Label(
        content_frame,
        text="✓ Analysis Completed Successfully!",
        style='Success.Header.TLabel',
        justify='left'
        ).pack(pady=(0, 15))
    
        # Output file label with dynamic wrapping
        file_frame = ttk.Frame(content_frame)
        file_frame.pack(fill='x', pady=(5, 10), anchor='w')
    
        # Special label for file path with word-breaking
        file_path_label = ttk.Label(
        file_frame,
        text=f"• Output File : {output_file}",
        style='Success.TLabel',
        wraplength=450  # Will wrap at 450 pixels
        )
        file_path_label.pack(side='left', fill='x', expand=True, anchor='w')
    
        # Runtime label
        ttk.Label(
        content_frame,
        text=f"• Execution Time : {runtime:.2f} seconds",
        style='Success.TLabel'
        ).pack(anchor='w', pady=(0, 15))
        
        # Action buttons
        btn_frame = ttk.Frame(content_frame)
        btn_frame.pack(pady=(15,0))
    
        open_file_btn = ttk.Button(
        btn_frame,
        text="Open File",
        command=lambda: self.open_file(output_file, win)
        )
        open_file_btn.pack(side='left')
        
        # Update window to calculate proper wrapping
        win.update_idletasks()
    
        # Adjust wrapping dynamically if needed
        width = getattr(file_path_label, 'winfo_reqwidth', lambda: 0)()
        if isinstance(width, int) and width > 450:
            file_path_label.config(wraplength=450)
    
        # Force window to front and maintain focus
        win.after(100, lambda: win.focus_force())
    
        # Add error handling for window close
        win.protocol("WM_DELETE_WINDOW", win.destroy)

    def _center_window(self, window):
        """Center any window on screen"""
        window.update_idletasks()
        width = window.winfo_width()
        height = window.winfo_height()
        x = (window.winfo_screenwidth() // 2) - (width // 2)
        y = (window.winfo_screenheight() // 2) - (height // 2)
        window.geometry(f'+{x}+{y}')
    

    def cleanup_after_analysis(self):
        """Safe cleanup – survives early window destruction."""
        try:
            if hasattr(self, 'progress') and self.progress:
                self.progress.stop()
        except tk.TclError:
            pass

        self.close_progress_window()   # re-use helper above

        try:
            self.proceed_button.config(state='normal')
        except tk.TclError:
            pass   # main window already dead
            
    def execute_analysis(self, input_excel_file, regex_input=None, positions_to_highlight=None, conservation_threshold=60, domain_params=None):
        """OPTIMIZED analysis function - writes directly with openpyxl"""
        print("\n" + "="*60)
        print("PROTEIN SEQUENCE ANALYSIS - TIMING BREAKDOWN")
        print("="*60)
        
        overall_start = time.time()
        
        # Detect file type
        file_ext = os.path.splitext(input_excel_file)[1].lower()
        is_fasta = file_ext in ['.fasta', '.fa', '.faa']
        
        print("\n[1/7] Reading and preprocessing data...")
        if is_fasta:
            print(f"  📄 Detected FASTA format: {os.path.basename(input_excel_file)}")
            df = read_fasta_file(input_excel_file)
        else:
            print(f"  📄 Detected Excel format: {os.path.basename(input_excel_file)}")
            df = read_and_preprocess_data(input_excel_file)
        
        print("[2/7] Cleaning and processing sequences...")
        process_start = time.time()
        df = clean_and_process_sequences(df)
        df_final, leading_columns_count = prepare_final_dataframe(df)
        process_time = time.time() - process_start
        print(f"  ⏱️  Sequence processing time: {process_time:.3f} seconds")
        
        # Create output file path
        output_dir = os.path.dirname(input_excel_file)
        output_file = os.path.join(output_dir, "msa_data_processed.xlsx")
        
        # OPTIMIZATION: Write directly with openpyxl, skip xlsxwriter
        print("[3/7] Creating workbook and writing data...")
        write_start = time.time()
        wb = Workbook()
        ws = wb.active
        ws.title = "First_Sheet"
        
        # Write data directly from dataframe
        for r_idx, row in enumerate(dataframe_to_rows(df_final, index=False, header=True), 1):
            ws.append(row)
        
        write_time = time.time() - write_start
        print(f"  ⏱️  Direct workbook creation time: {write_time:.3f} seconds")

        # Apply initial styles
        print("[4/7] Applying initial styles...")
        style_start = time.time()
        styles = setup_excel_styles()

        # APPLY HEADER STYLING (MISSING IN VERSION 2)
        for cell in ws[1]:
            cell.font = styles['bold_font']
            cell.alignment = styles['center_alignment']

        # Gene Name/ID columns left-aligned
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=leading_columns_count):
            for cell in row:
                cell.alignment = styles['center_alignment']

        adjust_column_widths(ws)
        apply_bulk_alignment(ws, styles['center_alignment'], min_row=2, min_col=leading_columns_count+1)
        
        style_time = time.time() - style_start
        print(f"  ⏱️  Initial styling time: {style_time:.3f} seconds")

        # NEW: collected motif hits (gapless coords) for optional MDCS calculation
        motif_hits_all = []

        
        # Process motif searching if selected
        if regex_input is not None:
            print("[5/7] Processing motif searching...")
            pattern_label_pairs = [(pattern, "") for pattern in regex_input]
    
            n_seq = len(df)
            # SMART THRESHOLD: Only use parallel if beneficial
            if n_seq < 100:
                n_procs = 1  # Serial is faster for small datasets
                print("  ℹ️  Using serial mode (dataset too small for parallel benefit)")
            else:
                n_procs = min(
                int(self.motif_parallel_processes.get()),
                max(1, multiprocessing.cpu_count() // 2),
                8
                )
            
            highlight_ranges, print_ranges, match_summary, pattern_print_ranges, highlight_ranges_freq, motif_hits_all = \
            apply_pattern_highlights(
            ws, df, pattern_label_pairs,
            leading_columns_count, styles,
            n_procs,
            conservation_thresh=conservation_threshold)
            
            
            # Apply frequent pattern borders
            border_start = time.time()
            total_rows = len(df['protein sequences_cleaned'])  # ✅ Use the correct column name
            threshold = (conservation_threshold / 100) * total_rows

           
            red_border = styles['red_thick_border']
 
            # Pre-filter conserved columns
            conserved_cols = {
               (start_col, end_col) 
               for (start_col, end_col), freq in           highlight_ranges_freq.items() 
               if freq >= threshold
            }

            if conserved_cols:
           # Batch apply borders column-wise
              for start_col, end_col in conserved_cols:
                for col in range(start_col, end_col + 1):
                  col_letter = get_column_letter(col)
                  
                  for row in range(2, ws.max_row + 1):
                      ws[f"{col_letter}{row}"].border = red_border

            border_time = time.time() - border_start
            print(f"  ⏱️  Conservation border application time: {border_time:.3f} seconds")
        else:
            print("[5/7] Motif searching skipped (not selected)")
        
        # Process position highlighting if selected
        if positions_to_highlight is not None:
            print("[6/7] Processing position highlighting...")
            apply_position_highlights(ws, positions_to_highlight, leading_columns_count, styles)
        else:
            print("[6/7] Position highlighting skipped (not selected)")
        
        # Create summary sheet if motif searching was done
        if regex_input is not None:
            print("Creating match summary sheet...")
            create_summary_sheet(wb, match_summary, pattern_print_ranges, total_rows, styles, pattern_label_pairs)

        # Process domain scanning if selected
        if domain_params is not None:
            print("[7/7] Running domain scanning with HMMER...")
            num_proc = domain_params.get('num_processes')        # cpu_cores already in dict
            if num_proc > 1:
                res_lines = run_parallel_domain_scanning(
                    df,
                    domain_params['pfam_db_path'],
                    domain_params['evalue_threshold'],
                    num_proc
                )
            else:
                res_lines = run_domain_scanning(
                    df,
                    domain_params['pfam_db_path'],
                    domain_params['evalue_threshold']
                )
            
            if res_lines:
                print(f"Found {len(res_lines)} domain hits. Creating domain highlights...")
                apply_domain_highlights(wb, ws, df, res_lines, leading_columns_count, styles, df_final)

                # NEW: MDCS Summary sheet (only when both motif searching and domain scanning are enabled)
                if regex_input is not None and len(motif_hits_all) > 0:
                    try:
                        create_mdcs_summary_sheet(wb, motif_hits_all, res_lines, styles)
                        print("Finished creating 'MDCS Summary' sheet.")
                    except Exception as e:
                        print(f"MDCS Summary creation skipped due to error: {e}")

                # Write hitdata.txt file
                hitdata_file = os.path.join(output_dir, "hitdata.txt")
                with open(hitdata_file, "w") as out_file:
                    out_file.write("\n".join(res_lines))
                print(f"Domain hits written to {hitdata_file}")
            else:
                print("No domain hits found with current E-value threshold.")
        else:
            print("[7/7] Domain scanning skipped (not selected)")
        
        # Save workbook
        print("\nSaving final workbook...")
        save_start = time.time()
        wb.save(output_file)
        save_time = time.time() - save_start
        print(f"  ⏱️  Final workbook save time: {save_time:.3f} seconds")
        
        overall_time = time.time() - overall_start
        
        print("\n" + "="*60)
        print("TIMING SUMMARY")
        print("="*60)
        print(f"Total execution time: {overall_time:.3f} seconds")
        print("="*60 + "\n")
        
        print(f"Analysis complete. Output saved to: {output_file}")
        return output_file

def main():
    root = tk.Tk()
    app = ProteinAnalysisApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
