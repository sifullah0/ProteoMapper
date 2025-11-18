"""
Unit tests for core protein sequence analysis functions.
"""

import pytest
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import tempfile
import os


from ProteoMapper import (
    detect_header_row,
    read_and_preprocess_data,
    clean_and_process_sequences,
    prepare_final_dataframe,
    load_patterns,
    setup_excel_styles,
    adjust_column_widths,
    apply_bulk_alignment,
    load_highlight_positions,
    HeaderNotFoundError,
    FileReadError,
    RegexFileError,
    PositionsFileError
)


class TestHeaderDetection:
    """Test suite for header row detection in Excel files."""
    
    def test_when_valid_header_exists_expect_correct_index(self):
        """Test detection of valid header with all required columns."""
        df = pd.DataFrame([
            ['Some data', 'Other data'],
            ['Protein Sequences', 'Gene Name'],
            ['ACDEFG', 'GeneA']
        ])
        
        result = detect_header_row(df)
        assert result == 1, "Should detect header at index 1"
    
    def test_when_header_has_extra_spaces_expect_detection(self):
        """Test header detection with irregular spacing."""
        df = pd.DataFrame([
            ['  Protein   Sequences  ', '  Gene   Name  '],
            ['ACDEFG', 'GeneA']
        ])
        
        result = detect_header_row(df)
        assert result == 0, "Should normalize spaces and detect header"
    
    def test_when_gene_id_instead_of_gene_name_expect_detection(self):
        """Test header detection with Gene ID column."""
        df = pd.DataFrame([
            ['Protein Sequences', 'Gene ID'],
            ['ACDEFG', 'Gene001']
        ])
        
        result = detect_header_row(df)
        assert result == 0, "Should detect header with Gene ID"
    
    def test_when_header_missing_expect_none(self):
        """Test behavior when required columns are absent."""
        df = pd.DataFrame([
            ['Random Column', 'Another Column'],
            ['Data1', 'Data2']
        ])
        
        result = detect_header_row(df)
        assert result is None, "Should return None for missing header"
    
    def test_when_empty_dataframe_expect_none(self):
        """Test handling of empty DataFrame."""
        df = pd.DataFrame()
        
        result = detect_header_row(df)
        assert result is None, "Should return None for empty DataFrame"
    
    def test_when_case_insensitive_header_expect_detection(self):
        """Test case-insensitive header detection."""
        df = pd.DataFrame([
            ['PROTEIN SEQUENCES', 'GENE NAME'],
            ['ACDEFG', 'GeneA']
        ])
        
        result = detect_header_row(df)
        assert result == 0, "Should detect header regardless of case"


class TestSequenceCleaning:
    """Test suite for protein sequence cleaning and processing."""
    
    def test_when_valid_sequences_expect_proper_cleaning(self):
        """Test cleaning of valid protein sequences."""
        df = pd.DataFrame({
            'protein sequences': ['ACDEFGHIKLMNPQRSTVWY', 'MKKL-YFDS']
        })
        
        result = clean_and_process_sequences(df)
        
        assert 'protein sequences_cleaned' in result.columns
        assert result['protein sequences_cleaned'].iloc[0] == 'ACDEFGHIKLMNPQRSTVWY'
        assert result['protein sequences_cleaned'].iloc[1] == 'MKKL-YFDS'
    
    def test_when_sequences_have_leading_gt_expect_removal(self):
        """Test removal of leading '>' characters."""
        df = pd.DataFrame({
            'protein sequences': ['>ACDEFG', 'KLMNOP']
        })
        
        result = clean_and_process_sequences(df)
        
        assert result['has_leading_gt'].iloc[0] == True
        assert result['has_leading_gt'].iloc[1] == False
        assert result['Protein sequences'].iloc[0] == 'ACDEFG'
    
    def test_when_sequences_have_invalid_chars_expect_removal(self):
        """Test removal of invalid amino acid characters."""
        df = pd.DataFrame({
            'protein sequences': ['ACX123EFG', 'KLM*NOP!']
        })
        
        result = clean_and_process_sequences(df)
        
        # Invalid chars removed but gaps preserved in cleaned
        assert 'X' not in result['protein sequences_cleaned'].iloc[0]
        assert '1' not in result['protein sequences_cleaned'].iloc[0]
        
        # Match version removes all non-standard
        assert '*' not in result['protein sequences_match'].iloc[1]
        assert '!' not in result['protein sequences_match'].iloc[1]
    
    def test_when_empty_sequences_expect_empty_strings(self):
        """Test handling of empty or NaN sequences."""
        df = pd.DataFrame({
            'protein sequences': ['', np.nan, 'ACDEFG']
        })
        
        result = clean_and_process_sequences(df)
        
        assert result['protein sequences_cleaned'].iloc[0] == ''
        assert result['protein sequences_cleaned'].iloc[1] == ''
        assert result['protein sequences_cleaned'].iloc[2] == 'ACDEFG'
    
    def test_when_very_long_sequence_expect_complete_processing(self):
        """Test processing of very long sequences (>10000 AA)."""
        long_seq = 'ACDEFGHIKLMNPQRSTVWY' * 500  # 10000 characters
        df = pd.DataFrame({
            'protein sequences': [long_seq]
        })
        
        result = clean_and_process_sequences(df)
        
        assert len(result['protein sequences_cleaned'].iloc[0]) == 10000
        assert all(c in 'ACDEFGHIKLMNPQRSTVWY' for c in result['protein sequences_match'].iloc[0])
    
    def test_when_gaps_in_sequence_expect_preservation_in_cleaned(self):
        """Test that gaps are preserved in cleaned version."""
        df = pd.DataFrame({
            'protein sequences': ['AC--DEFG', 'KL-MN-OP']
        })
        
        result = clean_and_process_sequences(df)
        
        assert '--' in result['protein sequences_cleaned'].iloc[0]
        assert '-' in result['protein sequences_cleaned'].iloc[1]
        
        # Match version removes gaps
        assert '-' not in result['protein sequences_match'].iloc[0]


class TestDataFramePreparation:
    """Test suite for final DataFrame preparation."""
    
    def test_when_sequences_split_expect_correct_columns(self):
        """Test sequence splitting into individual columns."""
        df = pd.DataFrame({
            'gene name': ['GeneA', 'GeneB'],
            'protein sequences_cleaned': ['ACDE', 'FGHI'],
            'protein sequences_match': ['ACDE', 'FGHI']
        })
        
        df_final, leading_cols = prepare_final_dataframe(df)
        
        assert leading_cols == 1  # Only Gene Name
        assert '1' in df_final.columns
        assert '4' in df_final.columns
        assert df_final['1'].iloc[0] == 'A'
        assert df_final['4'].iloc[1] == 'I'
    
    def test_when_both_gene_name_and_id_expect_both_columns(self):
        """Test retention of both Gene Name and Gene ID."""
        df = pd.DataFrame({
            'gene name': ['GeneA'],
            'gene id': ['ID001'],
            'protein sequences_cleaned': ['ACDE'],
            'protein sequences_match': ['ACDE']
        })
        
        df_final, leading_cols = prepare_final_dataframe(df)
        
        assert leading_cols == 2
        assert 'Gene Name' in df_final.columns
        assert 'Gene ID' in df_final.columns
    
    def test_when_empty_gene_columns_expect_removal(self):
        """Test removal of empty gene columns."""
        df = pd.DataFrame({
            'gene name': ['', '', ''],
            'gene id': ['ID001', 'ID002', 'ID003'],
            'protein sequences_cleaned': ['ACE', 'DFG', 'HIK'],
            'protein sequences_match': ['ACE', 'DFG', 'HIK']
        })
        
        df_final, leading_cols = prepare_final_dataframe(df)
        
        assert 'Gene Name' not in df_final.columns
        assert 'Gene ID' in df_final.columns
        assert leading_cols == 1
    
    def test_when_unequal_sequence_lengths_expect_padding(self):
        """Test handling of sequences with different lengths."""
        df = pd.DataFrame({
            'gene name': ['GeneA', 'GeneB'],
            'protein sequences_cleaned': ['AC', 'DEFGHI'],
            'protein sequences_match': ['AC', 'DEFGHI']
        })
        
        df_final, _ = prepare_final_dataframe(df)
        
        # Should create 6 columns for longest sequence
        assert '6' in df_final.columns
        # Shorter sequence should have NaN in extra columns
        assert pd.isna(df_final['3'].iloc[0])


class TestPatternLoading:
    """Test suite for regex pattern loading from files."""
    
    def test_when_valid_patterns_file_expect_correct_parsing(self, tmp_path):
        """Test loading of valid regex patterns."""
        pattern_file = tmp_path / "patterns.txt"
        pattern_file.write_text("ACE::Motif1\nDFG::Motif2\nHIK")
        
        result = load_patterns(str(pattern_file))
        
        assert len(result) == 3
        assert result[0] == ('ACE', 'Motif1')
        assert result[1] == ('DFG', 'Motif2')
        assert result[2] == ('HIK', '')
    
    def test_when_empty_lines_in_file_expect_skip(self, tmp_path):
        """Test skipping of empty lines."""
        pattern_file = tmp_path / "patterns.txt"
        pattern_file.write_text("ACE::Motif1\n\n\nDFG::Motif2")
        
        result = load_patterns(str(pattern_file))
        
        assert len(result) == 2
    
    def test_when_whitespace_in_patterns_expect_trimming(self, tmp_path):
        """Test trimming of whitespace in patterns and labels."""
        pattern_file = tmp_path / "patterns.txt"
        pattern_file.write_text("  ACE  ::  Motif1  \n  DFG::Motif2")
        
        result = load_patterns(str(pattern_file))
        
        assert result[0] == ('ACE', 'Motif1')
        assert result[1] == ('DFG', 'Motif2')
    
    def test_when_file_not_found_expect_exception(self):
        """Test handling of missing pattern file."""
        with pytest.raises(RegexFileError):
            load_patterns("nonexistent_file.txt")
    
    def test_when_complex_regex_patterns_expect_preservation(self, tmp_path):
        """Test preservation of complex regex patterns."""
        pattern_file = tmp_path / "patterns.txt"
        pattern_file.write_text("[AC]{2,4}E::Complex\n(DFG|HIK)+::Alternative")
        
        result = load_patterns(str(pattern_file))
        
        assert result[0] == ('[AC]{2,4}E', 'Complex')
        assert result[1] == ('(DFG|HIK)+', 'Alternative')


class TestPositionLoading:
    """Test suite for position highlighting file loading."""
    
    def test_when_valid_positions_expect_correct_parsing(self, tmp_path):
        """Test loading of valid position numbers."""
        pos_file = tmp_path / "positions.txt"
        pos_file.write_text("10 20 30 45 67")
        
        result = load_highlight_positions(str(pos_file))
        
        assert result == [10, 20, 30, 45, 67]
    
    def test_when_positions_on_multiple_lines_expect_parsing(self, tmp_path):
        """Test parsing positions across multiple lines."""
        pos_file = tmp_path / "positions.txt"
        pos_file.write_text("10 20\n30 45\n67")
        
        result = load_highlight_positions(str(pos_file))
        
        assert result == [10, 20, 30, 45, 67]
    
    def test_when_non_numeric_values_expect_skip(self, tmp_path):
        """Test skipping of non-numeric values."""
        pos_file = tmp_path / "positions.txt"
        pos_file.write_text("10 abc 20 xyz 30")
        
        result = load_highlight_positions(str(pos_file))
        
        assert result == [10, 20, 30]
    
    def test_when_file_not_found_expect_exception(self):
        """Test handling of missing positions file."""
        with pytest.raises(PositionsFileError):
            load_highlight_positions("nonexistent_file.txt")
    
    def test_when_empty_file_expect_empty_list(self, tmp_path):
        """Test handling of empty positions file."""
        pos_file = tmp_path / "positions.txt"
        pos_file.write_text("")
        
        result = load_highlight_positions(str(pos_file))
        
        assert result == []


class TestExcelStyling:
    """Test suite for Excel styling operations."""
    
    def test_when_setup_styles_expect_all_style_objects(self):
        """Test creation of all required style objects."""
        styles = setup_excel_styles()
        
        assert 'sky_blue_fill' in styles
        assert 'green_fill' in styles
        assert 'orange_fill' in styles
        assert 'red_thick_border' in styles
        assert 'bold_font' in styles
        assert 'center_alignment' in styles
        assert 'left_alignment' in styles
        
        # Verify types
        assert isinstance(styles['sky_blue_fill'], PatternFill)

    @pytest.mark.parametrize("cols,expected_leading", [
        (["gene name"],            ["Gene Name"]),
        (["gene id"],              ["Gene ID"]),
        (["gene name", "gene id"], ["Gene Name", "Gene ID"]),
        ([],                       []),
    ])
    def test_gene_column_variations(self, cols, expected_leading):
        base = {"protein sequences_cleaned": ["ACD"], "protein sequences_match": ["ACD"]}
        for c in cols:
            base[c] = ["dummy"]
        df = pd.DataFrame(base)

        df_final, leading_cols = prepare_final_dataframe(df)

        assert df_final.columns[:len(expected_leading)].tolist() == expected_leading
        assert leading_cols == len(expected_leading)
    
    def test_when_adjust_column_widths_expect_proper_sizing(self):
        """Test column width adjustment based on content."""
        wb = Workbook()
        ws = wb.active

        # Add some data
        ws.append(['Gene Name', 'A', 'C', 'D', 'E'])
        ws.append(['GeneA', 'A', 'C', 'D', 'E'])

        adjust_column_widths(ws)

        # Verify that width is assigned (not None) and is reasonable
        for col in ('A', 'C', 'D', 'E'):
            assert ws.column_dimensions[col].width is not None, \
                f"Column {col} should have a width assigned"
            assert ws.column_dimensions[col].width > 0, \
                f"Column {col} width should be positive"
        
        # First column should be wider than single-character columns
        assert ws.column_dimensions['A'].width > ws.column_dimensions['C'].width
        
        # Test max width capping at 50
        ws['A1'] = 'X' * 100  # Very long string
        adjust_column_widths(ws)
        assert ws.column_dimensions['A'].width <= 50, \
            "Width should be capped at 50"
    
    def test_when_bulk_alignment_applied_expect_all_cells_aligned(self):
        """Test bulk alignment application."""
        wb = Workbook()
        ws = wb.active
        styles = setup_excel_styles()
        
        # Add data
        for i in range(5):
            ws.append(['Data'] * 3)
        
        apply_bulk_alignment(ws, styles['center_alignment'], min_row=2, max_row=5)
        
        # Check that cells are center-aligned
        assert ws['A2'].alignment.horizontal == 'center'
        assert ws['C5'].alignment.horizontal == 'center'


class TestEdgeCases:
    """Test suite for edge cases and boundary conditions."""
    
    def test_when_single_sequence_expect_processing(self):
        """Test processing of single sequence."""
        df = pd.DataFrame({
            'protein sequences': ['ACDEFG']
        })
        
        result = clean_and_process_sequences(df)
        assert len(result) == 1
    
    def test_when_all_gaps_sequence_expect_handling(self):
        """Test sequence consisting only of gaps."""
        df = pd.DataFrame({
            'protein sequences': ['-----']
        })
        
        result = clean_and_process_sequences(df)
        assert result['protein sequences_cleaned'].iloc[0] == '-----'
        assert result['protein sequences_match'].iloc[0] == ''
    
    def test_when_unicode_characters_expect_removal(self):
        """Test handling of Unicode characters."""
        df = pd.DataFrame({
            'protein sequences': ['ACDE™FG§HI']
        })
        
        result = clean_and_process_sequences(df)
        # Unicode chars should be removed
        assert '™' not in result['protein sequences_cleaned'].iloc[0]
        assert '§' not in result['protein sequences_cleaned'].iloc[0]
    
        # In tests/test_core_functions.py
    def test_when_lowercase_sequences_expect_uppercase_in_match(self):
        """Test case handling in sequences."""
        df = pd.DataFrame({
        'protein sequences': ['acdefg']
        })

        result = clean_and_process_sequences(df)
    
        # Original case preserved in cleaned → lowercase
        assert result['protein sequences_cleaned'].iloc[0] == 'acdefg'
    
        # But match column is uppercase (for regex)
        assert result['protein sequences_match'].iloc[0] == 'ACDEFG'


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
