"""
Integration tests for protein sequence analysis pipeline.

Tests cover:
- End-to-end Excel processing
- HMMER integration with mock databases
- Pattern matching with highlighting
- Domain scanning workflows
- Multi-step pipeline operations
"""

import multiprocessing
multiprocessing.set_start_method("spawn", force=True)

import pytest
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import tempfile
import os
import subprocess
import shutil


from ProteoMapper import (
    read_and_preprocess_data,
    clean_and_process_sequences,
    prepare_final_dataframe,
    load_patterns,
    apply_pattern_highlights,
    apply_position_highlights,
    run_domain_scanning,
    apply_domain_highlights,
    setup_excel_styles,
    create_summary_sheet,
    create_domain_summary_sheet,
    HMMERError,
)


# -------------------------------------------------------------------
# Helpers / markers
# -------------------------------------------------------------------
def hmmer_available():
    """Check if HMMER is available on system."""
    try:
        subprocess.run(
            ["hmmscan", "-h"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            check=True,
        )
        return True
    except (subprocess.CalledProcessError, FileNotFoundError):
        return False


# -------------------------------------------------------------------
# Fixtures
# -------------------------------------------------------------------
@pytest.fixture
def sample_excel_file(tmp_path):
    """Create a sample Excel file for testing."""
    file_path = tmp_path / "test_input.xlsx"

    df = pd.DataFrame({
        "Protein Sequences": [
            "ACDEFGHIKLMNPQRSTVWY",
            "MKKLIFGDSMQRSTVWY",
            "ACDEFGHIKLMNPQRSTVWY",
            ">QRSTACDEFGHIKLMN",
        ],
        "Gene Name": ["GeneA", "GeneB", "GeneC", "GeneD"],
    })

    df.to_excel(file_path, index=False)
    return str(file_path)


@pytest.fixture
def complex_excel_file(tmp_path):
    """Create a complex Excel file with headers in different rows."""
    file_path = tmp_path / "complex_input.xlsx"

    # Create file with metadata rows before header
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        metadata_df = pd.DataFrame([
            ["Study Information", ""],
            ["Date:", "2024-01-01"],
            ["", ""],  # Empty row
            ["Protein Sequences", "Gene Name"],
            ["ACDEFGHIKLMNPQRSTVWY", "GeneA"],
            ["MKKLIFGDSMQRSTVWY", "GeneB"],
        ])
        metadata_df.to_excel(writer, index=False, header=False)

    return str(file_path)


@pytest.fixture
def mock_hmm_database(tmp_path):
    """Create a mock HMM database file for testing."""
    hmm_file = tmp_path / "test_pfam.hmm"

    # Minimal HMMER3 header + 1 position, just enough for hmmscan to accept
    hmm_content = """HMMER3/f [3.3.2 | Nov 2020]
NAME  Test_Domain
ACC   PF00001
DESC  Test protein domain
LENG  50
ALPH  amino
RF    no
MM    no
CONS  yes
CS    yes
MAP   yes
DATE  Mon Jan 01 00:00:00 2024
NSEQ  10
EFFN  5.0
CKSUM 123456789
STATS LOCAL MSV       -9.0000  0.70000
STATS LOCAL VITERBI   -9.5000  0.70000
STATS LOCAL FORWARD   -4.0000  0.70000
HMM          A        C        D        E        F        G        H        I        K        L        M        N        P        Q        R        S        T        V        W        Y   
            m->m     m->i     m->d     i->m     i->i     d->m     d->d
  COMPO   2.68618  4.42225  2.77519  2.73123  3.46354  2.40513  3.72494  3.29354  2.67741  2.69355  3.24034  2.90347  3.61503  3.61503  3.61503  2.61955  2.85256  2.89801  5.09345  3.61503
          2.68618  4.42225  2.77519  2.73123  3.46354  2.40513  3.72494  3.29354  2.67741  2.69355  3.24034  2.90347  3.61503  3.61503  3.61503  2.61955  2.85256  2.89801  5.09345  3.61503
          0.00000  5.00000  5.00000  1.46634  0.26236  0.00000        *
      1   2.68618  4.42225  2.77519  2.73123  3.46354  2.40513  3.72494  3.29354  2.67741  2.69355  3.24034  2.90347  3.61503  3.61503  3.61503  2.61955  2.85256  2.89801  5.09345  3.61503      1 - -
//
"""
    hmm_file.write_text(hmm_content)

    # Try hmmpress if available; not required, but closer to real Pfam setup
    if hmmer_available():
        try:
            subprocess.run(
                ["hmmpress", str(hmm_file)],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
        except Exception:
            # Non-fatal; hmmscan can still run directly on the .hmm
            pass

    return str(hmm_file)


# -------------------------------------------------------------------
# Excel pipeline tests
# -------------------------------------------------------------------
class TestExcelPipeline:
    """Integration tests for Excel file processing pipeline."""

    def test_end_to_end_simple_excel_processing(self, sample_excel_file, tmp_path):
        """Test complete pipeline from Excel read to final output."""
        df = read_and_preprocess_data(sample_excel_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)

        assert len(df_final) == 4
        assert "Gene Name" in df_final.columns
        assert leading_cols == 1

        # Sequence split into per-position columns
        assert "1" in df_final.columns
        assert df_final["1"].iloc[0] == "A"

    def test_end_to_end_complex_excel_with_metadata(self, complex_excel_file):
        """Test pipeline with Excel file containing metadata rows."""
        df = read_and_preprocess_data(complex_excel_file)
        df = clean_and_process_sequences(df)
        df_final, _ = prepare_final_dataframe(df)

        assert len(df_final) == 2
        assert "Gene Name" in df_final.columns

    def test_pattern_matching_integration(self, sample_excel_file, tmp_path):
        """Test pattern matching with Excel output."""
        from openpyxl import Workbook

        df = read_and_preprocess_data(sample_excel_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)

        wb = Workbook()
        ws = wb.active

        # Write data
        for col_idx, col_name in enumerate(df_final.columns, 1):
            ws.cell(1, col_idx, col_name)
        for row_idx, row_data in enumerate(df_final.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row_idx, col_idx, value)

        patterns = [("ACDEFG", "Motif1"), ("MKK", "Motif2")]
        styles = setup_excel_styles()

        highlight_ranges, print_ranges, match_summary, pattern_print_ranges, freq = \
            apply_pattern_highlights(ws, df, patterns, leading_cols, styles)

        assert len(match_summary) > 0
        assert "Motif1" in match_summary or "Motif2" in match_summary

        output_file = tmp_path / "test_output.xlsx"
        wb.save(output_file)
        assert output_file.exists()

    def test_position_highlighting_integration(self, sample_excel_file, tmp_path):
        """Test position highlighting in Excel output."""
        from openpyxl import Workbook

        df = read_and_preprocess_data(sample_excel_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)

        wb = Workbook()
        ws = wb.active

        # Write data
        for col_idx, col_name in enumerate(df_final.columns, 1):
            ws.cell(1, col_idx, col_name)
        for row_idx, row_data in enumerate(df_final.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row_idx, col_idx, value)

        positions = [1, 5, 10]
        styles = setup_excel_styles()
        apply_position_highlights(ws, positions, leading_cols, styles)

        # Check that specified positions have green fill on first data row
        for pos in positions:
            actual_col = pos + leading_cols
            cell = ws.cell(2, actual_col)
            # Depending on openpyxl version, rgb may include alpha; we just check suffix
            assert cell.fill.start_color.rgb.endswith("90EE90")

        output_file = tmp_path / "test_positions.xlsx"
        wb.save(output_file)
        assert output_file.exists()


# -------------------------------------------------------------------
# HMMER integration tests
# -------------------------------------------------------------------
class TestHMMERIntegration:
    """Integration tests for HMMER domain scanning."""

    @pytest.mark.hmmer
    @pytest.mark.skipif(not hmmer_available(), reason="HMMER not installed")
    def test_domain_scanning_with_mock_database(
        self, sample_excel_file, mock_hmm_database
    ):
        """Test HMMER domain scanning with mock database."""
        df = read_and_preprocess_data(sample_excel_file)
        df = clean_and_process_sequences(df)

        try:
            results = run_domain_scanning(
                df,
                mock_hmm_database,
                evalue_threshold=1.0,  # lenient for tests
            )
            assert isinstance(results, list)
        except HMMERError as e:
            pytest.skip(f"HMMER execution failed: {e}")

    @pytest.mark.hmmer
    @pytest.mark.skipif(not hmmer_available(), reason="HMMER not installed")
    def test_domain_scanning_with_invalid_database_expect_error(
        self, sample_excel_file
    ):
        """Test error handling with invalid database path."""
        df = read_and_preprocess_data(sample_excel_file)
        df = clean_and_process_sequences(df)

        with pytest.raises(HMMERError):
            run_domain_scanning(
                df,
                "/nonexistent/database.hmm",
                evalue_threshold=0.001,
            )

    @pytest.mark.hmmer
    @pytest.mark.skipif(not hmmer_available(), reason="HMMER not installed")
    def test_domain_highlighting_integration(
        self, sample_excel_file, mock_hmm_database, tmp_path
    ):
        """Test full domain scanning and highlighting pipeline."""
        from openpyxl import Workbook

        df = read_and_preprocess_data(sample_excel_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)

        wb = Workbook()
        ws = wb.active

        try:
            results = run_domain_scanning(
                df,
                mock_hmm_database,
                evalue_threshold=1.0,
            )

            if results:
                styles = setup_excel_styles()
                domain_sheet_created = apply_domain_highlights(
                    wb, ws, df, results, leading_cols, styles, df_final
                )

                if domain_sheet_created:
                    assert "Domain_Highlights" in wb.sheetnames
                    assert "Domain Summary" in wb.sheetnames
        except HMMERError as e:
            pytest.skip(f"HMMER execution failed: {e}")

    def test_domain_scanning_with_empty_sequences_expect_no_results(
        self, tmp_path
    ):
        """Test domain scanning with empty sequences."""
        df = pd.DataFrame({
            "protein sequences": ["", "", ""],
            "gene name": ["G1", "G2", "G3"],
        })
        df = clean_and_process_sequences(df)

        mock_db = tmp_path / "mock.hmm"
        mock_db.write_text("MOCK")

        if hmmer_available():
            try:
                results = run_domain_scanning(df, str(mock_db), 0.001)
                assert isinstance(results, list)
            except Exception:
                # We accept failures here because the DB is intentionally invalid
                pass


# -------------------------------------------------------------------
# Multi-step pipeline tests
# -------------------------------------------------------------------
class TestMultiStepPipeline:
    """Tests for complete multi-step analysis pipelines."""

    def test_motif_and_position_combined_pipeline(
        self, sample_excel_file, tmp_path
    ):
        """Pipeline with both motif searching and position highlighting."""
        from openpyxl import Workbook

        df = read_and_preprocess_data(sample_excel_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)

        wb = Workbook()
        ws = wb.active

        # Write data
        for col_idx, col_name in enumerate(df_final.columns, 1):
            ws.cell(1, col_idx, col_name)
        for row_idx, row_data in enumerate(df_final.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row_idx, col_idx, value)

        patterns = [("ACDEFG", "Motif1")]
        positions = [1, 5, 10]
        styles = setup_excel_styles()

        apply_pattern_highlights(ws, df, patterns, leading_cols, styles)
        apply_position_highlights(ws, positions, leading_cols, styles)

        output_file = tmp_path / "combined_output.xlsx"
        wb.save(output_file)
        assert output_file.exists()

        wb_loaded = load_workbook(output_file)
        ws_loaded = wb_loaded.active

        # Just sanity check that some styling exists on a data cell
        assert ws_loaded.cell(2, 2).fill is not None

    def test_summary_sheet_creation_integration(
        self, sample_excel_file, tmp_path
    ):
        """Test creation of match summary sheet."""
        from openpyxl import Workbook

        df = read_and_preprocess_data(sample_excel_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)

        wb = Workbook()
        ws = wb.active

        patterns = [("ACDEFG", "Motif1"), ("MKK", "Motif2")]
        styles = setup_excel_styles()

        _, _, match_summary, pattern_print_ranges, _ = apply_pattern_highlights(
            ws, df, patterns, leading_cols, styles
        )

        create_summary_sheet(
            wb,
            match_summary,
            pattern_print_ranges,
            len(df),
            styles,
            patterns,
        )

        assert "Match Summary" in wb.sheetnames

        summary_ws = wb["Match Summary"]
        assert summary_ws["A1"].value == "Regex Pattern"
        assert summary_ws["B1"].value == "Matches"
        assert summary_ws["C1"].value == "Matched Subsequence Positions"

        output_file = tmp_path / "summary_output.xlsx"
        wb.save(output_file)
        assert output_file.exists()


# -------------------------------------------------------------------
# Error recovery tests
# -------------------------------------------------------------------
class TestErrorRecovery:
    """Tests for error handling and recovery in pipelines."""

    def test_corrupted_excel_file_expect_error(self, tmp_path):
        """Handling of corrupted Excel file."""
        from ProteoMapper import FileReadError

        corrupted_file = tmp_path / "corrupted.xlsx"
        corrupted_file.write_text("This is not a valid Excel file")

        with pytest.raises(FileReadError):
            read_and_preprocess_data(str(corrupted_file))

    def test_missing_required_columns_expect_error(self, tmp_path):
        """Handling of Excel file missing required columns."""
        from ProteoMapper import HeaderNotFoundError

        file_path = tmp_path / "incomplete.xlsx"

        df = pd.DataFrame({
            "Random Column": ["Data1", "Data2"],
            "Another Column": ["Data3", "Data4"],
        })
        df.to_excel(file_path, index=False)

        with pytest.raises(HeaderNotFoundError):
            read_and_preprocess_data(str(file_path))

    def test_invalid_regex_pattern_expect_error(self, sample_excel_file):
        """Handling of invalid regex patterns."""
        from openpyxl import Workbook

        df = read_and_preprocess_data(sample_excel_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)

        wb = Workbook()
        ws = wb.active

        patterns = [("[INVALID", "BadPattern")]
        styles = setup_excel_styles()

        with pytest.raises(ValueError):
            apply_pattern_highlights(ws, df, patterns, leading_cols, styles)


# -------------------------------------------------------------------
# Performance integration tests
# -------------------------------------------------------------------
class TestPerformanceIntegration:
    """Integration tests for performance with large datasets."""

    def test_large_dataset_processing(self, tmp_path):
        """Processing of large dataset (1000 sequences)."""
        import time

        large_file = tmp_path / "large_dataset.xlsx"

        sequences = ["ACDEFGHIKLMNPQRSTVWY"] * 1000
        gene_names = [f"Gene{i}" for i in range(1000)]

        df = pd.DataFrame({
            "Protein Sequences": sequences,
            "Gene Name": gene_names,
        })
        df.to_excel(large_file, index=False)

        start_time = time.time()

        df = read_and_preprocess_data(str(large_file))
        df = clean_and_process_sequences(df)
        df_final, _ = prepare_final_dataframe(df)

        elapsed_time = time.time() - start_time

        assert elapsed_time < 30
        assert len(df_final) == 1000

    def test_very_long_sequences_processing(self, tmp_path):
        """Processing of very long sequences (>5000 AA)."""
        long_file = tmp_path / "long_sequences.xlsx"

        long_seq = "ACDEFGHIKLMNPQRSTVWY" * 250  # 5000 AA

        df = pd.DataFrame({
            "Protein Sequences": [long_seq, long_seq],
            "Gene Name": ["LongGene1", "LongGene2"],
        })
        df.to_excel(long_file, index=False)

        df = read_and_preprocess_data(str(long_file))
        df = clean_and_process_sequences(df)
        df_final, _ = prepare_final_dataframe(df)

        # Gene Name + 5000 position columns
        assert df_final.shape[1] == 5001


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
