# tests/test_real_e2e.py
import pytest
import tkinter as tk
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


from ProteoMapper import ProteinAnalysisApp

@pytest.fixture
def app():
    root = tk.Tk()
    root.withdraw()
    app = ProteinAnalysisApp(root)
    yield app
    try:
        root.update_idletasks()
        root.destroy()
    except Exception:
        pass


@pytest.fixture
def sample_excel(tmp_path):
    file_path = tmp_path / "input.xlsx"
    df = pd.DataFrame({
        "Protein Sequences": [
            "ACDEFGHIKLMNPQRSTVWY",
            "ACDEFGHIKLMNPQRSTVWY",
            "MKKLIFGDSMQRSTVWY",
            "QRSTACDEFGHIKLMN"
        ],
        "Gene Name": ["GeneA", "GeneB", "GeneC", "GeneD"]
    })
    df.to_excel(file_path, index=False)
    return str(file_path)


# ------------------------------------------------------------------
#  1.  Motif-search end-to-end
# ------------------------------------------------------------------
def test_real_motif_search_end_to_end(app, sample_excel, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    output_file = tmp_path / "msa_data_processed.xlsx"

    # configure GUI
    app.motif_file_path.set(sample_excel)
    app.motif_search_var.set(True)
    app.regex_text.insert("1.0", "ACDEFG::CommonMotif\nMKK.*WY::FlexMotif")
    app.root.update_idletasks()

    # run analysis synchronously  –  use 40 % so 2/4 hits trigger border
    app.execute_analysis(
        input_excel_file=sample_excel,
        regex_input=["ACDEFG", "MKK.*WY"],
        positions_to_highlight=None,
        conservation_threshold=40,      # <<<< changed
        domain_params=None
    )

    # validation
    assert output_file.is_file()
    wb = load_workbook(output_file)
    ws = wb.active

    # sky-blue fill for motif  (RGB: 00BFFF)
    assert ws["B2"].fill.start_color.rgb == "0000BFFF"
    # thick red border (conservation)
    assert ws["B2"].border.left.style == "thick"
    # summary sheet created
    assert "Match Summary" in wb.sheetnames


# ------------------------------------------------------------------
#  2.  Position-highlighting end-to-end
# ------------------------------------------------------------------
def test_real_position_highlighting_end_to_end(app, sample_excel, tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    output_file = tmp_path / "msa_data_processed.xlsx"

    # populate GUI
    app.motif_file_path.set(sample_excel)
    app.position_highlight_var.set(True)
    app.positions_to_highlight.set("1 5 10")
    app.root.update_idletasks()

    # run analysis synchronously
    app.execute_analysis(
        input_excel_file=sample_excel,
        regex_input=None,
        positions_to_highlight=[1, 5, 10],
        conservation_threshold=60,
        domain_params=None
    )

    # validation
    assert output_file.is_file()
    wb = load_workbook(output_file)
    ws = wb.active

    # green fill for positions 1, 5, 10  -> columns B, F, K
    assert ws["B2"].fill.start_color.rgb == "0090EE90"   # pos 1
    assert ws["F2"].fill.start_color.rgb == "0090EE90"   # pos 5
    assert ws["K2"].fill.start_color.rgb == "0090EE90"   # pos 10