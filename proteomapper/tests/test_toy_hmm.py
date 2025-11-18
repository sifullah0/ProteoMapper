"""
CI-friendly HMMER tests that use a hand-made 50-aa model.

Coverage goal: every line inside
    run_domain_scanning()
    run_parallel_domain_scanning()
    apply_domain_highlights()
    create_domain_summary_sheet()
without downloading Pfam-A.
"""

import pytest
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess
from unittest.mock import Mock
import multiprocessing
multiprocessing.set_start_method("spawn", force=True)


from ProteoMapper import (
    read_and_preprocess_data,
    clean_and_process_sequences,
    prepare_final_dataframe,
    run_domain_scanning,
    run_parallel_domain_scanning,
    apply_domain_highlights,
    create_domain_summary_sheet,
    setup_excel_styles,
    HMMERError,
)

# ------------------------------------------------------------------
# Fixtures
# ------------------------------------------------------------------
@pytest.fixture
def toy_hmm_path():
    """Absolute path to the toy HMM (already hmm-pressed)."""
    return Path(__file__).with_name("fixtures") / "toy.hmm"


@pytest.fixture
def small_excel(tmp_path, toy_hmm_path):
    """Excel with one perfect-match sequence (guaranteed hit)."""
    perfect = (toy_hmm_path.parent / "perfect_seq.txt").read_text().strip()
    df = pd.DataFrame({
        "Protein Sequences": [perfect, "ACDEFGHIKLMNPQRSTVWY", ""],
        "Gene Name": ["Perfect", "GeneB", "Empty"],
    })
    file_ = tmp_path / "small.xlsx"
    df.to_excel(file_, index=False)
    return str(file_)


# ------------------------------------------------------------------
# Serial domain scanning tests
# ------------------------------------------------------------------
def test_domain_scanning_hits(toy_hmm_path, small_excel):
    """Exercise the full serial scan → parse → filter pipeline."""
    df = read_and_preprocess_data(small_excel)
    df = clean_and_process_sequences(df)

    hits = run_domain_scanning(
        df, str(toy_hmm_path), evalue_threshold=10
    )

    # Basic type & non-emptiness
    assert isinstance(hits, list), "Expected list of hits"
    assert len(hits) >= 1, f"Expected at least 1 hit from perfect_seq, got {len(hits)}"

    # Check output format (stable public contract)
    assert all("Sequence:" in h for h in hits), "Missing 'Sequence:' in hit"
    assert all("E-value:" in h for h in hits), "Missing 'E-value:' in hit"
    assert all("Start:" in h for h in hits), "Missing 'Start:' in hit"
    assert all("End:" in h for h in hits), "Missing 'End:' in hit"

    # For the hand-made fixture, first hit should be for sequence 1
    first_hit = hits[0]
    assert "Sequence:1" in first_hit, "First hit should be from sequence 1"


# ------------------------------------------------------------------
# Parallel domain scanning tests
# ------------------------------------------------------------------
def test_parallel_domain_scanning_hits(toy_hmm_path, small_excel):
    """
    Parallel scan returns well-formed hits and does not lose any domains,
    compared to the serial implementation.
    """
    df = read_and_preprocess_data(small_excel)
    df = clean_and_process_sequences(df)

    serial_hits = run_domain_scanning(df, str(toy_hmm_path), evalue_threshold=10)
    parallel_hits = run_parallel_domain_scanning(
        df, str(toy_hmm_path), evalue_threshold=10, num_processes=2
    )

    assert isinstance(parallel_hits, list), "Parallel hits should be a list"
    assert len(parallel_hits) >= 1, "Expected at least one parallel hit"

    # We expect every sequence to be scanned exactly once in both modes
    assert len(parallel_hits) == len(serial_hits), (
        "Parallel scanning should not change the total number of domain hits"
    )

    # Format of each parallel hit must match the public string contract
    assert all("Sequence:" in h for h in parallel_hits), "Missing 'Sequence:' in parallel hit"
    assert all("E-value:" in h for h in parallel_hits), "Missing 'E-value:' in parallel hit"
    assert all("Start:" in h for h in parallel_hits), "Missing 'Start:' in parallel hit"
    assert all("End:" in h for h in parallel_hits), "Missing 'End:' in parallel hit"


# ------------------------------------------------------------------
# Highlighting + summary tests (serial hits)
# ------------------------------------------------------------------
def test_domain_highlighting_sheet_creation(toy_hmm_path, small_excel, tmp_path):
    """End-to-end: Excel → serial scan → highlight → summary sheets."""
    from openpyxl import Workbook

    df = read_and_preprocess_data(small_excel)
    df = clean_and_process_sequences(df)
    df_final, leading_cols = prepare_final_dataframe(df)

    wb = Workbook()
    ws = wb.active
    # write header + data
    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    hits = run_domain_scanning(df, str(toy_hmm_path), 10)
    styles = setup_excel_styles()

    created = apply_domain_highlights(
        wb, ws, df, hits, leading_cols, styles, df_final
    )

    assert created is True, "apply_domain_highlights should return True when hits exist"
    assert "Domain_Highlights" in wb.sheetnames, "Missing Domain_Highlights sheet"
    assert "Domain Summary" in wb.sheetnames, "Missing Domain Summary sheet"

    # Save and sanity-check
    out = tmp_path / "domain_serial.xlsx"
    wb.save(out)

    wb2 = load_workbook(out)

    # Check Domain_Highlights sheet
    dom_ws = wb2["Domain_Highlights"]
    assert dom_ws.max_row >= 2, "Domain_Highlights should have header + data rows"
    assert dom_ws["A1"].value in ["Gene Name", "Gene ID"], "Missing header in Domain_Highlights"

    # Check Domain Summary sheet
    summary = wb2["Domain Summary"]
    assert summary.max_row >= 2, "Domain Summary should have header + data rows"

    headers = [cell.value for cell in summary[1]]
    expected_headers = [
        "Sequence", "Start(without '-')", "End(without '-')",
        "Domain Name", "Accession ID", "E-value", "Cond. E-value", "Bit Score",
    ]
    assert headers == expected_headers, f"Unexpected headers: {headers}"


def test_parallel_highlighting_sheet_creation(toy_hmm_path, small_excel, tmp_path):
    """
    End-to-end: Excel → *parallel* scan → highlight → summary sheets.

    We don't assume identical Sequence IDs to the serial pipeline,
    only that domain hits can be highlighted and summarized.
    """
    from openpyxl import Workbook

    df = read_and_preprocess_data(small_excel)
    df = clean_and_process_sequences(df)
    df_final, leading_cols = prepare_final_dataframe(df)

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    hits = run_parallel_domain_scanning(df, str(toy_hmm_path), 10, num_processes=2)
    styles = setup_excel_styles()

    created = apply_domain_highlights(
        wb, ws, df, hits, leading_cols, styles, df_final
    )

    assert created is True, "apply_domain_highlights should return True when parallel hits exist"
    assert "Domain_Highlights" in wb.sheetnames, "Missing Domain_Highlights sheet (parallel)"
    assert "Domain Summary" in wb.sheetnames, "Missing Domain Summary sheet (parallel)"

    out = tmp_path / "domain_parallel.xlsx"
    wb.save(out)

    wb2 = load_workbook(out)
    dom_ws = wb2["Domain_Highlights"]
    assert dom_ws.max_row >= 2
    summary = wb2["Domain Summary"]
    assert summary.max_row >= 2


# ------------------------------------------------------------------
# Edge cases & error handling
# ------------------------------------------------------------------
def test_empty_sequences_graceful(toy_hmm_path):
    """No hits → empty list, no crash."""
    df = pd.DataFrame({
        "Protein Sequences": ["", "XXXX"],
        "Gene Name": ["E1", "E2"],
    })
    df = clean_and_process_sequences(df)

    hits = run_domain_scanning(df, str(toy_hmm_path), 0.001)
    assert hits == [], "Empty/invalid sequences should return empty hit list"


def test_hmmer_invalid_database(tmp_path):
    """Missing/invalid database file should raise HMMERError."""
    df = pd.DataFrame({
        "Protein Sequences": ["ACDEFGHIKLMNPQRSTVWY"],
        "Gene Name": ["TestGene"],
    })
    df = clean_and_process_sequences(df)

    fake_db = tmp_path / "does_not_exist.hmm"

    with pytest.raises(HMMERError, match="hmmscan failed"):
        run_domain_scanning(df, str(fake_db), 0.001)


def test_hmmer_execution_failure(toy_hmm_path, monkeypatch):
    """Simulate hmmscan returning non-zero exit code."""
    df = pd.DataFrame({
        "Protein Sequences": ["ACDEFGHIKLMNPQRSTVWY"],
        "Gene Name": ["TestGene"],
    })
    df = clean_and_process_sequences(df)

    def mock_failed_run(*args, **kwargs):
        result = Mock()
        result.returncode = 1
        result.stderr = "Error: hmmscan execution failed"
        result.stdout = ""
        return result

    monkeypatch.setattr(subprocess, "run", mock_failed_run)

    with pytest.raises(HMMERError, match="hmmscan failed with return code 1"):
        run_domain_scanning(df, str(toy_hmm_path), 0.001)


# ------------------------------------------------------------------
# Highlighting behavior with gapped sequences
# ------------------------------------------------------------------
def test_domain_highlights_with_gaps(toy_hmm_path, tmp_path):
    """Domain highlighting must handle gapped sequences correctly."""
    from openpyxl import Workbook

    perfect = (toy_hmm_path.parent / "perfect_seq.txt").read_text().strip()

    # Add gaps to simulate an MSA: insert '---' after every 10 residues
    gapped_seq = ""
    for i, char in enumerate(perfect):
        gapped_seq += char
        if (i + 1) % 10 == 0:
            gapped_seq += "---"

    df = pd.DataFrame({
        "Protein Sequences": [gapped_seq],
        "Gene Name": ["GappedSeq"],
    })
    df = clean_and_process_sequences(df)
    df_final, leading_cols = prepare_final_dataframe(df)

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    hits = run_domain_scanning(df, str(toy_hmm_path), evalue_threshold=10)

    if not hits:
        pytest.skip("No HMMER hits found - cannot test highlighting")

    styles = setup_excel_styles()
    created = apply_domain_highlights(wb, ws, df, hits, leading_cols, styles, df_final)

    assert created is True, "Should handle gapped sequences"
    assert "Domain_Highlights" in wb.sheetnames

    dom_ws = wb["Domain_Highlights"]

    # Look for any orange-filled cell (domain highlighting)
    orange_found = False
    cells_with_fill = []

    for row_idx in range(2, dom_ws.max_row + 1):
        for col_idx in range(leading_cols + 1, dom_ws.max_column + 1):
            cell = dom_ws.cell(row=row_idx, column=col_idx)

            if hasattr(cell.fill, "start_color"):
                color = cell.fill.start_color
                if hasattr(color, "index"):
                    color_value = color.index
                elif hasattr(color, "rgb"):
                    color_value = color.rgb
                else:
                    color_value = str(color)

                # Orange fill (depending on style / index representation)
                if color_value in ["FFA500", "00FFA500", "FFFFA500"]:
                    orange_found = True
                    cells_with_fill.append((row_idx, col_idx, color_value))

    assert orange_found, (
        f"Domain highlighting (orange) should be applied. "
        f"Found {len(cells_with_fill)} orange cells. "
        f"Total hits: {len(hits)}"
    )


def test_no_domain_hits_returns_false(tmp_path):
    """apply_domain_highlights returns False and doesn't create sheets when no hits."""
    from openpyxl import Workbook

    df = pd.DataFrame({
        "Protein Sequences": ["ACDEFG"],
        "Gene Name": ["ShortSeq"],
    })
    df = clean_and_process_sequences(df)
    df_final, leading_cols = prepare_final_dataframe(df)

    wb = Workbook()
    ws = wb.active
    for r in dataframe_to_rows(df_final, index=False, header=True):
        ws.append(r)

    styles = setup_excel_styles()

    created = apply_domain_highlights(wb, ws, df, [], leading_cols, styles, df_final)

    assert created is False, "Should return False when no domain hits"
    assert "Domain_Highlights" not in wb.sheetnames, "Should not create sheet without hits"
    assert "Domain Summary" not in wb.sheetnames, "Should not create summary without hits"


# ------------------------------------------------------------------
# E-value filtering semantics
# ------------------------------------------------------------------
def test_evalue_filtering(toy_hmm_path, small_excel):
    """Stricter E-value threshold should yield fewer or equal hits."""
    df = read_and_preprocess_data(small_excel)
    df = clean_and_process_sequences(df)

    # Very strict threshold - should get fewer hits
    strict_hits = run_domain_scanning(
        df, str(toy_hmm_path), evalue_threshold=1e-20
    )

    # Permissive threshold - should get more hits
    permissive_hits = run_domain_scanning(
        df, str(toy_hmm_path), evalue_threshold=10
    )

    assert len(strict_hits) <= len(permissive_hits), (
        "Stricter E-value threshold should yield fewer or equal hits"
    )

    # Verify all strict hits meet the threshold
    for hit in strict_hits:
        evalue_str = hit.split("E-value:")[1].split("|")[0].strip()
        evalue = float(evalue_str)
        assert evalue <= 1e-20, f"Hit E-value {evalue} exceeds threshold 1e-20"
