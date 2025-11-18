"""
Performance and stress tests for protein sequence analysis.

"""

import pytest
import pandas as pd
import numpy as np
import time
import tempfile
import os
import psutil
import gc
from openpyxl import Workbook


from ProteoMapper import (
    read_and_preprocess_data,
    clean_and_process_sequences,
    prepare_final_dataframe,
    apply_pattern_highlights,
    apply_position_highlights,
    setup_excel_styles
)


def get_memory_usage():
    """Get current memory usage in MB."""
    process = psutil.Process()
    return process.memory_info().rss / (1024 * 1024)


@pytest.fixture
def large_dataset_file(tmp_path):
    """Create large dataset for performance testing."""
    file_path = tmp_path / "large_dataset.xlsx"
    
    # Create 1000 sequences
    sequences = ['ACDEFGHIKLMNPQRSTVWY'] * 1000
    gene_names = [f'Gene{i:04d}' for i in range(1000)]
    
    df = pd.DataFrame({
        'Protein Sequences': sequences,
        'Gene Name': gene_names
    })
    
    df.to_excel(file_path, index=False)
    return str(file_path)


@pytest.fixture
def very_large_dataset_file(tmp_path):
    """Create very large dataset (5000 sequences)."""
    file_path = tmp_path / "very_large_dataset.xlsx"
    
    sequences = ['ACDEFGHIKLMNPQRSTVWY'] * 5000
    gene_names = [f'Gene{i:05d}' for i in range(5000)]
    
    df = pd.DataFrame({
        'Protein Sequences': sequences,
        'Gene Name': gene_names
    })
    
    df.to_excel(file_path, index=False)
    return str(file_path)


@pytest.fixture
def long_sequences_file(tmp_path):
    """Create file with very long sequences."""
    file_path = tmp_path / "long_sequences.xlsx"
    
    # Create 100 sequences of 5000 AA each
    long_seq = 'ACDEFGHIKLMNPQRSTVWY' * 250
    sequences = [long_seq] * 100
    gene_names = [f'LongGene{i:03d}' for i in range(100)]
    
    df = pd.DataFrame({
        'Protein Sequences': sequences,
        'Gene Name': gene_names
    })
    
    df.to_excel(file_path, index=False)
    return str(file_path)


class TestDataProcessingPerformance:
    """Performance tests for data processing functions."""
    
    @pytest.mark.performance
    def test_large_dataset_read_performance(self, large_dataset_file, benchmark):
        """Benchmark reading large Excel files."""
        def read_data():
            return read_and_preprocess_data(large_dataset_file)
        
        result = benchmark(read_data)
        assert len(result) == 1000
    
    @pytest.mark.performance
    def test_sequence_cleaning_performance(self, large_dataset_file, benchmark):
        """Benchmark sequence cleaning on large dataset."""
        df = read_and_preprocess_data(large_dataset_file)
        
        def clean_sequences():
            return clean_and_process_sequences(df.copy())
        
        result = benchmark(clean_sequences)
        assert 'protein sequences_cleaned' in result.columns
    
    @pytest.mark.performance
    def test_dataframe_preparation_performance(self, large_dataset_file, benchmark):
        """Benchmark DataFrame preparation."""
        df = read_and_preprocess_data(large_dataset_file)
        df = clean_and_process_sequences(df)
        
        def prepare_df():
            return prepare_final_dataframe(df.copy())
        
        result, _ = benchmark(prepare_df)
        assert len(result) == 1000
    
    @pytest.mark.performance
    def test_end_to_end_processing_time(self, large_dataset_file):
        """Test complete processing time for large dataset."""
        start_time = time.time()
        
        df = read_and_preprocess_data(large_dataset_file)
        df = clean_and_process_sequences(df)
        df_final, _ = prepare_final_dataframe(df)
        
        elapsed_time = time.time() - start_time
        
        # Should complete in under 30 seconds for 1000 sequences
        assert elapsed_time < 30.0
        assert len(df_final) == 1000
    
    @pytest.mark.slow
    @pytest.mark.performance
    def test_very_large_dataset_processing(self, very_large_dataset_file):
        """Test processing of very large dataset (5000 sequences)."""
        start_time = time.time()
        mem_start = get_memory_usage()
        
        df = read_and_preprocess_data(very_large_dataset_file)
        df = clean_and_process_sequences(df)
        df_final, _ = prepare_final_dataframe(df)
        
        elapsed_time = time.time() - start_time
        mem_end = get_memory_usage()
        mem_used = mem_end - mem_start
        
        # Should complete in under 2 minutes
        assert elapsed_time < 120.0
        assert len(df_final) == 5000
        
        # Memory usage should be reasonable (< 1GB increase)
        assert mem_used < 1024.0
        
        print(f"\nProcessing 5000 sequences:")
        print(f"  Time: {elapsed_time:.2f}s")
        print(f"  Memory: {mem_used:.2f}MB")


class TestPatternMatchingPerformance:
    """Performance tests for pattern matching operations."""
    
    @pytest.mark.performance
    def test_single_pattern_matching_performance(self, large_dataset_file, benchmark):
        """Benchmark single pattern matching on large dataset (serial path)."""
        df = read_and_preprocess_data(large_dataset_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)
        
        wb = Workbook()
        ws = wb.active
        
        patterns = [('ACDEFG', 'TestMotif')]
        styles = setup_excel_styles()
        
        # Keep this benchmark in serial mode to avoid multiprocess noise
        def match_patterns():
            return apply_pattern_highlights(
                ws, df, patterns, leading_cols, styles, n_procs=1
            )
        
        benchmark(match_patterns)
    
    @pytest.mark.performance
    def test_multiple_patterns_matching_performance(self, large_dataset_file):
        """
        Test performance with multiple regex patterns using parallel motif matching.

        This explicitly uses n_procs=2 to exercise the multiprocessing path in
        apply_pattern_highlights and _match_chunk, improving coverage.
        """
        df = read_and_preprocess_data(large_dataset_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)
        
        wb = Workbook()
        ws = wb.active
        
        # Test with 10 different patterns
        patterns = [
            ('ACDEFG', 'Motif1'),
            ('HIKLMN', 'Motif2'),
            ('PQRST', 'Motif3'),
            ('VWY', 'Motif4'),
            ('[AC]{3}', 'Motif5'),
            ('DEF.*HIK', 'Motif6'),
            ('LMN|PQR', 'Motif7'),
            ('K{2,}', 'Motif8'),
            ('ST[VW]', 'Motif9'),
            ('G.{2}H', 'Motif10')
        ]
        styles = setup_excel_styles()
        
        start_time = time.time()
        # n_procs=2 → trigger multiprocessing branch and _match_chunk
        _, _, match_summary, pattern_print_ranges, _ = apply_pattern_highlights(
            ws, df, patterns, leading_cols, styles, n_procs=2
        )
        elapsed_time = time.time() - start_time
        
        # Should complete in under 60 seconds for 1000 sequences x 10 patterns
        assert elapsed_time < 60.0
        
        # Basic sanity checks: parallel path actually found matches
        # At least some of the motifs should be present and have hits.
        assert isinstance(match_summary, dict) or hasattr(match_summary, "items")
        assert any(count > 0 for count in match_summary.values()), \
            "Parallel motif matching found no hits; expected at least one."
        
        # Motif1 ('ACDEFG') is guaranteed to occur in every test sequence
        assert match_summary.get('Motif1', 0) > 0
        
        print(f"\nMatching 10 patterns across 1000 sequences (parallel): {elapsed_time:.2f}s")
    
    @pytest.mark.performance
    def test_complex_regex_performance(self, large_dataset_file):
        """Test performance with complex regex patterns (still allowed to be serial)."""
        df = read_and_preprocess_data(large_dataset_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)
        
        wb = Workbook()
        ws = wb.active
        
        # Complex patterns
        patterns = [
            ('[ACDEFG]{5,10}', 'ComplexMotif1'),
            ('(HIK|LMN){2,}', 'ComplexMotif2'),
            ('P[QRST]{2,4}[VWY]', 'ComplexMotif3')
        ]
        styles = setup_excel_styles()
        
        start_time = time.time()
        apply_pattern_highlights(ws, df, patterns, leading_cols, styles, n_procs=1)
        elapsed_time = time.time() - start_time
        
        # Should handle complex patterns efficiently
        assert elapsed_time < 45.0



class TestExcelOperationsPerformance:
    """Performance tests for Excel I/O operations."""
    
    @pytest.mark.performance
    def test_excel_writing_performance(self, large_dataset_file, tmp_path):
        """Test performance of writing large dataset to Excel."""
        df = read_and_preprocess_data(large_dataset_file)
        df = clean_and_process_sequences(df)
        df_final, _ = prepare_final_dataframe(df)
        
        output_file = tmp_path / "perf_output.xlsx"
        
        start_time = time.time()
        
        wb = Workbook()
        ws = wb.active
        
        # Write data
        for col_idx, col_name in enumerate(df_final.columns, 1):
            ws.cell(1, col_idx, col_name)
        for row_idx, row_data in enumerate(df_final.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row_idx, col_idx, value)
        
        wb.save(output_file)
        
        elapsed_time = time.time() - start_time
        
        # Should complete in under 45 seconds
        assert elapsed_time < 45.0
        assert output_file.exists()
        
        print(f"\nWriting 1000 sequences to Excel: {elapsed_time:.2f}s")
    
    @pytest.mark.performance
    def test_styling_performance(self, large_dataset_file):
        """Test performance of applying styles to large workbook."""
        df = read_and_preprocess_data(large_dataset_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)
        
        wb = Workbook()
        ws = wb.active
        
        # Write minimal data
        for i in range(100):
            ws.append(['Data'] * 25)
        
        styles = setup_excel_styles()
        
        start_time = time.time()
        
        # Apply position highlighting (simpler than patterns)
        positions = list(range(1, 21))  # Highlight 20 positions
        apply_position_highlights(ws, positions, leading_cols, styles)
        
        elapsed_time = time.time() - start_time
        
        # Should be very fast (< 5 seconds)
        assert elapsed_time < 5.0


class TestMemoryEfficiency:
    """Tests for memory efficiency and leak detection."""
    
    @pytest.mark.performance
    def test_memory_usage_large_dataset(self, large_dataset_file):
        """Memory cleanup: strict when RSS grows, permissive when it doesn't."""
        import gc, psutil, os

        process = psutil.Process(os.getpid())
        MB = 1024 * 1024

        mem_start = process.memory_info().rss / MB

        # ---- work sequence ----
        df = read_and_preprocess_data(large_dataset_file)
        mem_read = process.memory_info().rss / MB

        df = clean_and_process_sequences(df)
        mem_clean = process.memory_info().rss / MB

        df_final, _ = prepare_final_dataframe(df)
        mem_prep = process.memory_info().rss / MB

        prep_increase = mem_prep - mem_clean

        # ---- cleanup ----
        del df, df_final
        gc.collect()
        mem_end = process.memory_info().rss / MB
        cleanup_decrease = mem_prep - mem_end   # how much was given back

        # ---- reporting ----
        print(f"\nMemory (MB)  read+{mem_read-mem_start:.2f}  "
            f"clean+{mem_clean-mem_read:.2f}  prep+{prep_increase:.2f}  "
            f"freed-{cleanup_decrease:.2f}")

        # ---- assertion ----
        if prep_increase > 5.0:                 # meaningful growth → strict
            assert cleanup_decrease > prep_increase * 0.7
            print(f"   Strict mode: {cleanup_decrease:.2f} MB freed (>70% of {prep_increase:.2f} MB)")
        else:                                   # tiny / zero → just don’t leak
            assert cleanup_decrease >= 0
            print(f"   Permissive mode: No leak detected (growth ≤5MB)")
    
    @pytest.mark.performance
    def test_memory_leak_repeated_processing(self, large_dataset_file):
        """Test for memory leaks with repeated processing."""
        gc.collect()
        mem_start = get_memory_usage()
        
        # Process multiple times
        for i in range(5):
            df = read_and_preprocess_data(large_dataset_file)
            df = clean_and_process_sequences(df)
            df_final, _ = prepare_final_dataframe(df)
            del df, df_final
            gc.collect()
        
        mem_end = get_memory_usage()
        mem_increase = mem_end - mem_start
        
        # Memory increase should be minimal (< 50MB total for 5 iterations)
        assert mem_increase < 50.0
        
        print(f"\nMemory increase after 5 iterations: {mem_increase:.2f}MB")


class TestScalability:
    """Tests for scalability with varying dataset sizes."""
    
    @pytest.mark.performance
    def test_scalability_with_increasing_sequences(self, tmp_path):
        """Test how processing time scales with number of sequences."""
        sizes = [100, 500, 1000, 2000]
        times = []
        
        for size in sizes:
            # Create dataset
            file_path = tmp_path / f"dataset_{size}.xlsx"
            df = pd.DataFrame({
                'Protein Sequences': ['ACDEFGHIKLMNPQRSTVWY'] * size,
                'Gene Name': [f'Gene{i}' for i in range(size)]
            })
            df.to_excel(file_path, index=False)
            
            # Time processing
            start_time = time.time()
            df = read_and_preprocess_data(str(file_path))
            df = clean_and_process_sequences(df)
            df_final, _ = prepare_final_dataframe(df)
            elapsed_time = time.time() - start_time
            
            times.append(elapsed_time)
            
            # Cleanup
            file_path.unlink()
        
        print(f"\nScalability test:")
        for size, time_val in zip(sizes, times):
            print(f"  {size:4d} sequences: {time_val:6.2f}s")
        
        # Time should scale roughly linearly (not exponentially)
        # Check that doubling size doesn't more than triple time
        if len(times) >= 2:
            ratio = times[-1] / times[0]
            size_ratio = sizes[-1] / sizes[0]
            assert ratio < (size_ratio * 1.5)
    
    @pytest.mark.slow
    @pytest.mark.performance
    def test_scalability_with_sequence_length(self, tmp_path):
        """Test how processing scales with sequence length."""
        lengths = [100, 500, 1000, 5000]
        times = []
        
        for length in lengths:
            # Create dataset with specific sequence length
            file_path = tmp_path / f"seq_length_{length}.xlsx"
            seq = 'ACDEFGHIKLMNPQRSTVWY' * (length // 20)
            
            df = pd.DataFrame({
                'Protein Sequences': [seq] * 100,
                'Gene Name': [f'Gene{i}' for i in range(100)]
            })
            df.to_excel(file_path, index=False)
            
            # Time processing
            start_time = time.time()
            df = read_and_preprocess_data(str(file_path))
            df = clean_and_process_sequences(df)
            df_final, _ = prepare_final_dataframe(df)
            elapsed_time = time.time() - start_time
            
            times.append(elapsed_time)
            
            # Cleanup
            file_path.unlink()
        
        print(f"\nSequence length scalability:")
        for length, time_val in zip(lengths, times):
            print(f"  {length:5d} AA: {time_val:6.2f}s")


class TestConcurrentPerformance:
    """Tests for performance under concurrent operations."""
    
    @pytest.mark.performance
    def test_memory_efficiency_with_concurrent_operations(self, large_dataset_file):
        """Test memory usage when multiple operations run."""
        gc.collect()
        mem_start = get_memory_usage()
        
        # Load data
        df = read_and_preprocess_data(large_dataset_file)
        df = clean_and_process_sequences(df)
        df_final, leading_cols = prepare_final_dataframe(df)
        
        # Create workbook and apply multiple operations
        wb = Workbook()
        ws = wb.active
        
        patterns = [('ACDEFG', 'Motif1'), ('HIKLMN', 'Motif2')]
        positions = [1, 5, 10, 15]
        styles = setup_excel_styles()
        
        # Apply both pattern matching and position highlighting
        apply_pattern_highlights(ws, df, patterns, leading_cols, styles)
        apply_position_highlights(ws, positions, leading_cols, styles)
        
        mem_peak = get_memory_usage()
        mem_used = mem_peak - mem_start
        
        # Memory should stay under 500MB for this workload
        assert mem_used < 500.0
        
        print(f"\nMemory used for concurrent operations: {mem_used:.2f}MB")


if __name__ == "__main__":
    # Run with: pytest test_performance.py -v -m performance
    pytest.main([__file__, "-v", "-m", "performance"])
